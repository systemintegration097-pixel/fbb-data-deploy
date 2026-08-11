import openpyxl

print("Analizando cobertura de 'Monthly Period' en reporte_bonus.xlsx...")
wb = openpyxl.load_workbook("reporte_bonus.xlsx", data_only=True, read_only=True)
sheet = wb.active

header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
monthly_col = next((i for i, c in enumerate(header) if c and "monthly" in str(c).lower()), None)
acc_col = next((i for i, c in enumerate(header) if c and "account" in str(c).lower()), 4)
print(f"Columna Monthly Period: {monthly_col} = '{header[monthly_col] if monthly_col is not None else 'NO ENCONTRADA'}'")
print(f"Columna Account: {acc_col}")
print()

months_found = set()
last_index = 0
total_rows = 0

for r in sheet.iter_rows(min_row=2, values_only=True):
    if not r:
        continue
    total_rows += 1
    if monthly_col is not None and len(r) > monthly_col and r[monthly_col]:
        months_found.add(str(r[monthly_col]).strip())
    if r[0]:  # INDEX column
        try:
            last_index = int(r[0])
        except:
            pass

wb.close()

print(f"Total filas leídas: {total_rows}")
print(f"Último INDEX() encontrado: {last_index}")
print(f"Meses cubiertos (Monthly Period):")
for m in sorted(months_found):
    print(f"  {m}")
}