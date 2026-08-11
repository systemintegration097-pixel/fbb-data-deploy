import openpyxl

print("--- REPORTE TABLEAU COLUMNS & SAMPLE ---")
try:
    wb_tab = openpyxl.load_workbook("reporte_tableau.xlsx", read_only=True)
    sheet = wb_tab.active
    rows = list(sheet.iter_rows(values_only=True))
    print("Total rows:", len(rows))
    if len(rows) > 0:
        print("Header (row 1):", rows[0])
    if len(rows) > 1:
        print("Sample row 2:", rows[1])
    if len(rows) > 2:
        print("Sample row 3:", rows[2])
except Exception as e:
    print("Error reading Tableau Excel:", e)

print("\n--- REPORTE GNOC COLUMNS & SAMPLE ---")
try:
    wb_gnoc = openpyxl.load_workbook("reporte_gnoc.xlsx", read_only=True)
    sheet_gnoc = wb_gnoc.active
    rows_gnoc = list(sheet_gnoc.iter_rows(values_only=True))
    print("Total rows:", len(rows_gnoc))
    if len(rows_gnoc) > 0:
        print("Header (row 1):", rows_gnoc[0])
    if len(rows_gnoc) > 1:
        print("Sample row 2:", rows_gnoc[1])
except Exception as e:
    print("Error reading GNOC Excel:", e)
