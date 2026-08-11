import os
import openpyxl

def inspect_openpyxl():
    path = r"c:\Users\jjvar\OneDrive\BITEL\Total de WO 18-12.xlsx"
    if not os.path.exists(path):
        print("File not found.")
        return
        
    print(f"Inspecting large Excel file: {path}")
    wb = openpyxl.load_workbook(path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        print("  Sheet is empty.")
        return
        
    print("  Columns:", list(header)[:10])
    
    create_idx = None
    for idx, col in enumerate(header):
        if col and 'create' in str(col).lower() and 'time' in str(col).lower():
            create_idx = idx
            break
            
    if create_idx is not None:
        print(f"  Found create time column at index {create_idx} ('{header[create_idx]}')")
        
        dates = []
        for r in rows:
            val = r[create_idx]
            if val:
                dates.append(str(val).strip())
                
        if dates:
            dates.sort()
            print("  Min Date (lexicographical):", dates[0])
            print("  Max Date (lexicographical):", dates[-1])
            
            import pandas as pd
            try:
                parsed = pd.to_datetime(dates, errors='coerce')
                print("  Chronological Min Date:", parsed.min())
                print("  Max Date:", parsed.max())
            except Exception as e:
                print("  Error parsing dates:", e)
        else:
            print("  No dates found.")
    else:
        print("  Create time column not found.")

if __name__ == '__main__':
    inspect_openpyxl()
