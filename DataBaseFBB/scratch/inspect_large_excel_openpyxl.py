import os
import openpyxl

def inspect_openpyxl():
    path = r"c:\Users\jjvar\OneDrive\BITEL\STATUS ENE-ABR WO.xlsx"
    if not os.path.exists(path):
        print("File not found.")
        return
        
    print(f"Inspecting large Excel file: {path}")
    wb = openpyxl.load_workbook(path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    
    target_sheets = ['GNOC', 'Data']
    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\nAnalyzing sheet: '{sheet_name}'")
            sheet = wb[sheet_name]
            
            # Read header row
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                print("  Sheet is empty.")
                continue
                
            print("  Columns:", list(header)[:10])
            
            # Find status and ft column indexes
            wo_status_idx = None
            ft_idx = None
            for idx, col in enumerate(header):
                if col:
                    col_lower = str(col).lower().strip()
                    if 'wo' in col_lower and 'status' in col_lower:
                        wo_status_idx = idx
                    elif col_lower == 'ft':
                        ft_idx = idx
                        
            if wo_status_idx is not None:
                print(f"  Found status column at index {wo_status_idx} ('{header[wo_status_idx]}')")
                if ft_idx is not None:
                    print(f"  Found FT column at index {ft_idx} ('{header[ft_idx]}')")
                else:
                    print("  FT column not found.")
                    
                total_rows = 0
                ft_inproc_count = 0
                ft_inproc_excl_count = 0
                
                # Count rows
                for r in rows:
                    total_rows += 1
                    status_val = r[wo_status_idx]
                    ft_val = r[ft_idx] if ft_idx is not None else None
                    
                    if status_val == 'FT Inprocessing':
                        ft_inproc_count += 1
                        if ft_val != 'vtp_marlo.delacruz':
                            ft_inproc_excl_count += 1
                            
                print(f"  Total data rows: {total_rows}")
                print(f"  Total FT Inprocessing: {ft_inproc_count}")
                print(f"  Total FT Inprocessing (excl. marlo.delacruz): {ft_inproc_excl_count}")
            else:
                print("  Status column not found in this sheet.")

if __name__ == '__main__':
    inspect_openpyxl()
