import openpyxl

def inspect_all():
    print("Cargando TODAS las WOs de gnoc sin filtro de fecha...")
    wb_gnoc = openpyxl.load_workbook("reporte_gnoc.xlsx", data_only=True)
    sheet_gnoc = wb_gnoc.active
    header_gnoc = [cell.value for cell in sheet_gnoc[8]]
    col_gnoc_wo = header_gnoc.index("WO code") if "WO code" in header_gnoc else 1
    col_gnoc_status = header_gnoc.index("WO Status") if "WO Status" in header_gnoc else 6
    col_gnoc_create = header_gnoc.index("Create Time") if "Create Time" in header_gnoc else 11
    
    all_gnoc_wos = {}
    for r in sheet_gnoc.iter_rows(min_row=9, values_only=True):
        if r and len(r) > col_gnoc_wo and r[col_gnoc_wo]:
            wo = str(r[col_gnoc_wo]).strip()
            status = str(r[col_gnoc_status]).strip() if len(r) > col_gnoc_status and r[col_gnoc_status] else ""
            cdate = str(r[col_gnoc_create]).strip() if len(r) > col_gnoc_create and r[col_gnoc_create] else ""
            all_gnoc_wos[wo] = {"status": status, "create": cdate}

    print(f"Total WOs totales en reporte_gnoc.xlsx: {len(all_gnoc_wos)}")

    wb_tab = openpyxl.load_workbook("reporte_tableau.xlsx", data_only=True)
    sheet_tab = wb_tab.active
    rows_tab = list(sheet_tab.iter_rows(values_only=True))
    
    found_count = 0
    not_found_count = 0
    
    sample_not_found = []
    
    for r in rows_tab[1:]:
        if r and len(r) > 2 and r[2]:
            wo = str(r[2]).strip()
            if wo in all_gnoc_wos:
                found_count += 1
            else:
                not_found_count += 1
                if len(sample_not_found) < 5:
                    sample_not_found.append((wo, r[4])) # WO_CODE, WO_CREATE_DATE

    print(f"\nCoincidencias totales en reporte_gnoc.xlsx sin filtro:")
    print(f"  Encontradas: {found_count}")
    print(f"  No encontradas en GNOC: {not_found_count}")
    print("Muestra de WOs en Tableau no encontradas en GNOC:")
    for wo, dt in sample_not_found:
        print(f"  WO: {wo}, Fecha Tableau: {dt}")

if __name__ == '__main__':
    inspect_all()
