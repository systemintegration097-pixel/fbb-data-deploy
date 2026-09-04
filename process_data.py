import os
import re
import glob
import sqlite3
import shutil
import openpyxl
from datetime import datetime
from nims_topology import parse_box_code

DB_PATH = "gnoc.db"
GNOC_EXCEL_PATH = "reporte_gnoc.xlsx"
# WOs 'Closed'/'Closed FT' cubren el mismo rango completo que las pendientes, pero
# descargadas en varios tramos (reporte_gnoc_closed_1.xlsx, _2.xlsx, ...) porque
# traerlas todas en una sola búsqueda satura el portal GNOC (ver download_report.py)
GNOC_CLOSED_EXCEL_GLOB = "reporte_gnoc_closed_*.xlsx"
TABLEAU_EXCEL_PATH = "reporte_tableau.xlsx"
BONUS_EXCEL_PATH = "reporte_bonus.xlsx"
# CNOC: sistema nuevo (ver download_cnoc.py) con WOs que no aparecen en GNOC. Mismo export
# (columnas con los mismos nombres que GNOC, en posiciones distintas) -> se reacomoda a la
# forma de GNOC en load_cnoc_rows() para reusar el mismo bucle de procesamiento de main().
CNOC_EXCEL_PATH = "reporte_cnoc.xlsx"
# Base del módulo FBB DATA (DataBaseFBB/), sincronizada aparte desde la pestaña "List of
# Boxes" del mismo Google Sheet que ya usa ese módulo -boxes.node_code es exactamente el
# mismo formato que el "Sub node code" de NIMS (ej. "ARE0005-XB01-HB01-SB111"), así que
# sirve para resolver el branch de una WO a partir de su caja cuando Tableau no lo trae.
FBB_DB_PATH = os.path.join("DataBaseFBB", "fbb_database.db")

def parse_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val
    date_str = str(date_val).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def get_branch_from_nims(unit_name, site_code):
    if not unit_name and not site_code:
        return ""
    
    # 1. Intentar con unit_name de NIMS
    if unit_name:
        u = unit_name.upper().strip()
        # Mapeos de Lima (ej: LIM-LI1BR, LI1-TEAM)
        if "LI1" in u: return "LI1"
        if "LI2" in u: return "LI2"
        if "LI3" in u: return "LI3"
        if "LI4" in u: return "LI4"
        if "LI7" in u: return "LI7"
        if "LI8" in u: return "LI8"
        # Mapeos generales de provincia por prefijo de 3 letras
        for b in ["ARE", "CAJ", "CUS", "HUN", "JUN", "LAL", "PIU", "SAN"]:
            if u.startswith(b):
                return b

    # 2. Intentar con site_code (ej: ARE0001 -> ARE)
    if site_code:
        s = site_code.upper().strip()
        for b in ["ARE", "CAJ", "CUS", "HUN", "JUN", "LAL", "PIU", "SAN"]:
            if s.startswith(b):
                return b
        # Lima sites (LIC, CAL, LIM, LI)
        if s.startswith("LIC") or s.startswith("CAL") or s.startswith("LIM") or s.startswith("LI"):
            return "LI4" # default para Lima si no se especifica equipo

    return ""

KNOWN_BRANCHES = {"ARE", "CAJ", "CUS", "HUN", "JUN", "LAL", "LI1", "LI2", "LI3", "LI4", "LI7", "LI8", "PIU", "SAN"}

def get_branch_from_responsible_unit(responsible_unit):
    """Último fallback: GNOC ya manda el branch embebido en responsible_unit/cd_group -no
    depende de NIMS en absoluto, así que resuelve el branch aunque el cliente sea tan nuevo
    que NIMS todavía no lo tenga (visto en producción: 9 WOs pendientes recientes sin branch
    ni connector_code porque la cuenta no existía aún en nims_subscribers).
    El formato varía según el equipo/team que lo genera en GNOC -visto en producción:
    'VTP_LI4BR FBB team', 'VTP_PIU - Team CANCHAQUE', etc.- así que en vez de exigir un patrón
    exacto completo, solo se toma el bloque alfanumérico pegado a 'VTP_'/'VTP-' y se valida
    contra KNOWN_BRANCHES (con o sin el sufijo 'BR' de "branch")."""
    if not responsible_unit:
        return ""
    m = re.match(r'^VTP[_-]([A-Z0-9]+)', responsible_unit.strip(), re.IGNORECASE)
    if not m:
        return ""
    code = m.group(1).upper()
    if code in KNOWN_BRANCHES:
        return code
    if code.endswith("BR") and code[:-2] in KNOWN_BRANCHES:
        return code[:-2]
    return ""

def load_boxes_branch_map():
    """Carga {node_code: branch} desde la tabla 'boxes' del módulo FBB DATA (sincronizada
    desde la pestaña 'List of Boxes' del Google Sheet). Es la fuente de verdad para
    resolver el branch de una WO a partir de su caja cuando Tableau no trae el branch."""
    if not os.path.exists(FBB_DB_PATH):
        print(f"[Aviso] No se encontró '{FBB_DB_PATH}'; no se podrá resolver branch por caja (List of Boxes).")
        return {}
    try:
        conn = sqlite3.connect(FBB_DB_PATH, timeout=30)
        cursor = conn.cursor()
        cursor.execute("SELECT node_code, branch FROM boxes WHERE node_code IS NOT NULL AND node_code != ''")
        boxes_map = {str(node_code).strip().upper(): (branch or "").strip() for node_code, branch in cursor.fetchall()}
        conn.close()
        print(f"  Cargadas {len(boxes_map)} cajas (List of Boxes) para resolución de branch.")
        return boxes_map
    except Exception as e:
        print(f"[Aviso] No se pudo cargar 'boxes' de '{FBB_DB_PATH}': {e}")
        return {}

def load_tableau_data(tableau_path):
    if not os.path.exists(tableau_path):
        print(f"[Aviso] No se encontró el archivo '{tableau_path}'. Las WOs no tendrán metadatos de Tableau.")
        return {}
        
    print(f"Abriendo y cargando datos de Tableau desde '{tableau_path}'...")
    tableau_map = {}
    try:
        wb = openpyxl.load_workbook(tableau_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            return {}
            
        header = rows[0]
        # Mapeo flexible de columnas por nombre
        indices = {
            "branch": 1, "wo_code": 2, "ticket_code": 3, "wo_create_date": 4,
            "responsible_unit_wo": 5, "ft_reassigned": 6, "account": 7,
            "warranty_period": 8, "implementation_test": 9, "act_status": 10,
            "sub_status": 11, "online_status": 12, "ft_gnoc": 13,
            "staf_team": 14, "connector_code": 15, "compcontent": 16
        }
        
        for idx, col_name in enumerate(header):
            if col_name:
                c = str(col_name).strip().upper()
                if "BRANCH" in c: indices["branch"] = idx
                elif "WO_CODE" in c or "WO CODE" in c: indices["wo_code"] = idx
                elif "TICKET_CODE" in c or "TICKET CODE" in c: indices["ticket_code"] = idx
                elif "WO_CREATE_DATE" in c: indices["wo_create_date"] = idx
                elif "RESPONSIBLE_UNIT" in c: indices["responsible_unit_wo"] = idx
                elif "ACCOUNT" in c: indices["account"] = idx
                elif "WARRANTY" in c: indices["warranty_period"] = idx
                elif "IMPLEMENTATION" in c: indices["implementation_test"] = idx
                elif "ACT STATUS" in c: indices["act_status"] = idx
                elif "SUB STATUS" in c: indices["sub_status"] = idx
                elif "ONLINE STATUS" in c: indices["online_status"] = idx
                elif "FT GNOC" in c: indices["ft_gnoc"] = idx
                elif "STAF TEAM" in c: indices["staf_team"] = idx
                elif "CONNECTOR" in c: indices["connector_code"] = idx
                elif "COMPCONTENT" in c: indices["compcontent"] = idx
                
        for r in rows[1:]:
            col_wo = indices["wo_code"]
            if r and len(r) > col_wo and r[col_wo]:
                wo_code = str(r[col_wo]).strip()
                def get_val(key):
                    i = indices[key]
                    return str(r[i]).strip() if len(r) > i and r[i] is not None else ""

                tableau_map[wo_code] = {
                    "branch": get_val("branch"),
                    "ticket_code": get_val("ticket_code"),
                    "wo_create_date": get_val("wo_create_date"),
                    "responsible_unit_wo": get_val("responsible_unit_wo"),
                    "account": get_val("account"),
                    "warranty_period": get_val("warranty_period"),
                    "implementation_test": get_val("implementation_test"),
                    "act_status": get_val("act_status"),
                    "sub_status": get_val("sub_status"),
                    "online_status": get_val("online_status"),
                    "ft_gnoc": get_val("ft_gnoc"),
                    "staf_team": get_val("staf_team"),
                    "connector_code": get_val("connector_code"),
                    "compcontent": get_val("compcontent")
                }
        wb.close()
        print(f"  Cargados {len(tableau_map)} registros de Tableau.")
    except Exception as e:
        print(f"Error al leer Tableau Excel: {e}")
        
    return tableau_map

def load_bonus_data(bonus_path):
    if not os.path.exists(bonus_path):
        print(f"[Aviso] No se encontró el archivo de Bonus '{bonus_path}'.")
        return {}
        
    print(f"Abriendo y cargando datos de Bonus desde '{bonus_path}'...")
    bonus_map = {}
    try:
        wb = openpyxl.load_workbook(bonus_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            wb.close()
            return {}
            
        header = rows[0]
        # Mapeo flexible de columnas por nombre. None = no encontrada en esta vista (en vez de
        # asumir una posición fija, que puede apuntar a la columna equivocada si esta fuente de
        # Tableau no trae ese dato, como pasa con 'staf_team' en la vista BonusCommistion actual).
        indices = {
            "account": 0,
            "implementation_test": 1,
            "staf_team": None
        }

        for idx, col_name in enumerate(header):
            if col_name:
                c = str(col_name).strip().upper()
                if "ACCOUNT" in c:
                    indices["account"] = idx
                elif "SECOND OF IMPLEMENTATION" in c or "IMPLEMENTATION" in c:
                    indices["implementation_test"] = idx
                elif "NAME_TECH_STAFF" in c or "STAF" in c or "LOWER" in c or "FT NAME" in c:
                    indices["staf_team"] = idx

        for r in rows[1:]:
            col_acc = indices["account"]
            if r and len(r) > col_acc and r[col_acc]:
                account = str(r[col_acc]).strip().lower()
                def get_val(key):
                    i = indices[key]
                    if i is None:
                        return ""
                    return str(r[i]).strip() if len(r) > i and r[i] is not None else ""
                
                bonus_map[account] = {
                    "implementation_test": get_val("implementation_test"),
                    "staf_team": get_val("staf_team")
                }
        wb.close()
        print(f"  Cargados {len(bonus_map)} registros de Bonus.")
    except Exception as e:
        print(f"Error al leer Bonus Excel: {e}")

    return bonus_map

# Columnas que GNOC y CNOC comparten con el mismo nombre de cabecera (posiciones distintas
# en cada export). Usado por load_cnoc_rows para reacomodar cada fila de CNOC al layout de
# columnas de GNOC y así reusar el mismo bucle de procesamiento de main() sin duplicarlo.
GNOC_CNOC_SHARED_COLUMNS = [
    "WO code", "WO Name", "WO type", "Description", "WO Status", "Create Time",
    "CD", "FT", "Priority", "FT comment", "System Code",
    "Closed Time(yyyy-MM-dd)", "FT completed time", "Subscribers",
]

def load_cnoc_rows(path, gnoc_header_len, gnoc_col_mapping, cutoff_date, header_row_idx):
    """Carga reporte_cnoc.xlsx (WOs de CNOC que no aparecen en GNOC, ver download_cnoc.py).
    Excluye WOs de prueba (Description contiene 'test', pedido explícito del usuario) y
    reacomoda cada fila al layout de columnas de GNOC -mismos nombres de columna, posiciones
    distintas- para poder reusar el mismo bucle de main() sin duplicar ~250 líneas de lógica."""
    if not os.path.exists(path):
        print(f"[Aviso] No se encontró el archivo de CNOC '{path}'; no se procesarán WOs de CNOC en esta sincronización.")
        return {}

    print(f"Abriendo archivo Excel de CNOC '{path}'...")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        cnoc_header = [cell.value for cell in sheet[header_row_idx]]
        cnoc_col_mapping = {col.strip(): idx for idx, col in enumerate(cnoc_header) if col}

        idx_wo_code = cnoc_col_mapping.get("WO code")
        idx_create_time = cnoc_col_mapping.get("Create Time")
        idx_description = cnoc_col_mapping.get("Description")
        if idx_wo_code is None or idx_create_time is None:
            print("  [Aviso] El archivo de CNOC no tiene las columnas esperadas ('WO code'/'Create Time'); se omite.")
            wb.close()
            return {}

        cnoc_by_wo_code = {}
        excluded_test = 0
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row_vals = list(row)
            if not row_vals or not any(v is not None for v in row_vals):
                continue
            if len(row_vals) <= idx_wo_code or len(row_vals) <= idx_create_time:
                continue

            wo_code = row_vals[idx_wo_code]
            if not wo_code:
                continue

            if idx_description is not None and len(row_vals) > idx_description:
                desc_val = row_vals[idx_description]
                if desc_val and "test" in str(desc_val).lower():
                    excluded_test += 1
                    continue

            create_dt = parse_date(row_vals[idx_create_time])
            if not create_dt or create_dt < cutoff_date:
                continue

            reshaped = [None] * gnoc_header_len
            for col_name in GNOC_CNOC_SHARED_COLUMNS:
                src_idx = cnoc_col_mapping.get(col_name)
                dst_idx = gnoc_col_mapping.get(col_name)
                if src_idx is not None and dst_idx is not None and src_idx < len(row_vals):
                    reshaped[dst_idx] = row_vals[src_idx]

            cnoc_by_wo_code[str(wo_code).strip()] = (reshaped, create_dt)

        wb.close()
        print(f"  Registros de CNOC leídos: {len(cnoc_by_wo_code)} (excluidas {excluded_test} WOs de prueba con 'test' en Description)")
        return cnoc_by_wo_code
    except Exception as e:
        print(f"Error al leer CNOC Excel: {e}")
        return {}

def main():
    if not os.path.exists(GNOC_EXCEL_PATH):
        print(f"Error: El archivo Excel '{GNOC_EXCEL_PATH}' no existe.")
        return

    print(f"Estableciendo conexión con la base de datos...")
    conn = sqlite3.connect(DB_PATH, timeout=30)
    cursor = conn.cursor()

    # 1. Cargar reglas de tipificación desde la base de datos
    cursor.execute("SELECT keyword, result FROM typification_rules")
    rules = cursor.fetchall()
    rules.sort(key=lambda x: len(x[0]), reverse=True)
    print(f"Cargadas {len(rules)} reglas de tipificación para análisis.")

    # 2. Cargar diccionario de Tableau y Bonus
    tableau_map = load_tableau_data(TABLEAU_EXCEL_PATH)
    bonus_map = load_bonus_data(BONUS_EXCEL_PATH)
    boxes_branch_map = load_boxes_branch_map()

    # 3. Leer Excel(s) de GNOC: el de pendientes (obligatorio) y el de cerradas (opcional,
    # solo existe si el downloader logró traer la ventana reciente de Closed/Closed FT)
    def open_gnoc_sheet(path, temp_suffix):
        print(f"Abriendo archivo Excel de GNOC '{path}'...")
        temp_excel_path = f"reporte_gnoc_temp_{temp_suffix}.xlsx"
        try:
            shutil.copy2(path, temp_excel_path)
            wb = openpyxl.load_workbook(temp_excel_path, data_only=True)
            return wb, wb.active, temp_excel_path
        except Exception as e:
            print(f"Advertencia al copiar temporal (cargando original directamente): {e}")
            wb = openpyxl.load_workbook(path, data_only=True)
            return wb, wb.active, None

    header_row_idx = 8
    cutoff_date = datetime(2025, 1, 1)

    def collect_rows(sheet, col_wo_code, col_create_time):
        rows = []
        latest_dt = None
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row_vals = list(row)
            if not row_vals or not any(v is not None for v in row_vals) or len(row_vals) <= max(col_wo_code, col_create_time):
                continue

            wo_code = row_vals[col_wo_code]
            if not wo_code:
                continue

            create_dt = parse_date(row_vals[col_create_time])
            if not create_dt or create_dt < cutoff_date:
                continue

            if latest_dt is None or create_dt > latest_dt:
                latest_dt = create_dt

            rows.append((row_vals, create_dt))
        return rows, latest_dt

    wb, sheet, temp_excel_path = open_gnoc_sheet(GNOC_EXCEL_PATH, "pending")
    header = [cell.value for cell in sheet[header_row_idx]]
    col_mapping = {col.strip(): idx for idx, col in enumerate(header) if col}

    # Columnas esperadas en GNOC (misma estructura de export para ambos archivos)
    col_wo_code = col_mapping.get("WO code", 1)
    col_wo_name = col_mapping.get("WO Name", 4)
    col_wo_type = col_mapping.get("WO type", 3)
    col_description = col_mapping.get("Description", 5)
    col_wo_status = col_mapping.get("WO Status", 6)
    col_create_time = col_mapping.get("Create Time", 11)
    col_cd = col_mapping.get("CD", 12)
    col_ft = col_mapping.get("FT", 13)
    col_priority = col_mapping.get("Priority", 14)
    col_ft_comment = col_mapping.get("FT comment", 25)
    col_system_code = col_mapping.get("System Code", 9)
    col_closed_time = col_mapping.get("Closed Time(yyyy-MM-dd)")
    col_ft_completed_time = col_mapping.get("FT completed time")
    col_subscribers = col_mapping.get("Subscribers")

    print("\nIniciando procesamiento de registros de GNOC...")

    pending_rows, latest_excel_date = collect_rows(sheet, col_wo_code, col_create_time)
    print(f"Registros de pendientes (FT Inprocessing): {len(pending_rows)}")
    pending_by_wo_code = {str(row_vals[col_wo_code]).strip(): (row_vals, create_dt) for row_vals, create_dt in pending_rows}

    closed_wbs = []
    closed_temp_paths = []
    closed_paths = sorted(glob.glob(GNOC_CLOSED_EXCEL_GLOB), key=lambda p: int(re.search(r'_(\d+)\.xlsx$', p).group(1)))
    closed_by_wo_code = {}
    if closed_paths:
        # Los tramos son acumulativos (cada búsqueda reexporta el rango completo hasta ese
        # punto, no un slice exclusivo de fechas): el último tramo ya contiene todo lo que
        # traen los anteriores. Cargar los 9 (cada uno con más filas casi-duplicadas) era
        # la causa real de que la sync tardara ~57 minutos: la mayor parte del tiempo se iba
        # en volver a parsear en memoria WOs ya vistas, no en trabajo nuevo. Basta con el
        # último tramo.
        last_closed_path = closed_paths[-1]
        closed_wb, closed_sheet, closed_temp_path = open_gnoc_sheet(last_closed_path, os.path.splitext(os.path.basename(last_closed_path))[0])
        closed_wbs.append(closed_wb)
        closed_temp_paths.append(closed_temp_path)
        closed_rows, _ = collect_rows(closed_sheet, col_wo_code, col_create_time)
        for row_vals, create_dt in closed_rows:
            closed_by_wo_code[str(row_vals[col_wo_code]).strip()] = (row_vals, create_dt)
        print(f"Registros de cerradas (Closed/Closed FT), usando el último tramo '{last_closed_path}' de {len(closed_paths)}: {len(closed_by_wo_code)} únicos")
    else:
        print(f"[Aviso] No se encontraron archivos '{GNOC_CLOSED_EXCEL_GLOB}'; no se procesarán WOs cerradas en esta sincronización.")

    # Una WO puede aparecer tanto en pendientes como en cerradas si se cerró entre que se
    # descargó cada reporte; nos quedamos con la versión "cerrada" por ser el estado final
    # (si insertáramos ambas, wo_code (UNIQUE) rompe el INSERT más abajo).
    combined_by_wo_code = dict(pending_by_wo_code)
    combined_by_wo_code.update(closed_by_wo_code)
    print(f"Total registros GNOC (únicos, pendientes + cerradas): {len(combined_by_wo_code)}")

    # CNOC: solo se agregan WOs que NO existan ya en GNOC (GNOC manda; CNOC solo llena huecos,
    # "hay WOs que no aparecen en GNOC").
    cnoc_by_wo_code = load_cnoc_rows(CNOC_EXCEL_PATH, len(header), col_mapping, cutoff_date, header_row_idx)
    cnoc_added = 0
    for wo_code_cnoc, val in cnoc_by_wo_code.items():
        if wo_code_cnoc not in combined_by_wo_code:
            combined_by_wo_code[wo_code_cnoc] = val
            cnoc_added += 1
    print(f"WOs de CNOC agregadas (no existían en GNOC): {cnoc_added} de {len(cnoc_by_wo_code)} leídas")

    rows_to_process = list(combined_by_wo_code.values())
    print(f"Total registros a procesar (GNOC + CNOC): {len(rows_to_process)}")

    now_dt = datetime.now()
    if latest_excel_date and latest_excel_date > now_dt:
        now_dt = latest_excel_date
        
    print(f"Hora de referencia para cálculo de pendientes: {now_dt.strftime('%d/%m/%Y %H:%M:%S')}")

    processed_data = []
    marlo_errors = 0
    classified_count = 0
    closed_without_comment = 0
    open_pending = 0
    matched_tableau_count = 0

    try:
        for row_vals, create_dt in rows_to_process:
            wo_code = str(row_vals[col_wo_code]).strip()
            wo_name = str(row_vals[col_wo_name]).strip() if row_vals[col_wo_name] else ""
            wo_type = str(row_vals[col_wo_type]).strip() if row_vals[col_wo_type] else ""
            description = str(row_vals[col_description]).strip() if row_vals[col_description] else ""
            wo_status = str(row_vals[col_wo_status]).strip() if row_vals[col_wo_status] else ""
            cd_group = str(row_vals[col_cd]).strip() if row_vals[col_cd] else ""
            ft_technician = str(row_vals[col_ft]).strip() if row_vals[col_ft] else ""
            priority = str(row_vals[col_priority]).strip() if row_vals[col_priority] else ""
            ft_comment = str(row_vals[col_ft_comment]).strip() if row_vals[col_ft_comment] else ""

            # GNOC ya trae la cuenta del abonado en su propia columna "Subscribers";
            # es la fuente más confiable (no depende de matchear Tableau ni de parsear
            # la descripción con regex), así que se usa como account primario más abajo.
            account_gnoc = ""
            if col_subscribers is not None and len(row_vals) > col_subscribers and row_vals[col_subscribers]:
                account_gnoc = str(row_vals[col_subscribers]).strip().lower()

            # "ft completed" es el equivalente de CNOC a "cerrada" (confirmado por el usuario:
            # en CNOC solo "FT Inprocessing"/"Pending" son pendientes reales; el resto,
            # incluido "FT completed", es para tipificación, no para matchear con Tableau).
            is_closed = wo_status.lower() in ("close", "closed", "closed ft", "ft completed")

            if is_closed:
                pending_hours = 0.0
            else:
                pending_hours = (now_dt - create_dt).total_seconds() / 3600.0
                if pending_hours < 0:
                    pending_hours = 0.0

            # Fecha/hora real de cierre y horas de resolución (solo aplica a WOs cerradas)
            closed_time_str = ""
            resolution_hours = None
            if is_closed:
                closed_time_raw = row_vals[col_closed_time] if col_closed_time is not None and len(row_vals) > col_closed_time else None
                if not closed_time_raw and col_ft_completed_time is not None and len(row_vals) > col_ft_completed_time:
                    closed_time_raw = row_vals[col_ft_completed_time]
                closed_dt = parse_date(closed_time_raw)
                if closed_dt:
                    closed_time_str = closed_dt.strftime("%Y-%m-%d %H:%M:%S")
                    resolution_hours = (closed_dt - create_dt).total_seconds() / 3600.0
                    if resolution_hours < 0:
                        resolution_hours = 0.0

            # Reglas especiales
            is_error = 0
            close_reason = "Pendiente"

            # 1. Regla especial para Marlo de la Cruz
            if "vtp_marlo.delacruz" in ft_technician.lower():
                is_error = 1
                close_reason = "Error (vtp_marlo.delacruz)"
                marlo_errors += 1
            else:
                # 2. Mapear comentario a su clasificación
                if ft_comment:
                    comment_lower = ft_comment.lower()
                    matched = False
                    for keyword, result in rules:
                        if keyword.lower() in comment_lower:
                            close_reason = result
                            matched = True
                            classified_count += 1
                            break
                    if not matched:
                        close_reason = "Sin clasificar"
                else:
                    if is_closed:
                        close_reason = "Cerrada sin comentario"
                        closed_without_comment += 1
                    else:
                        close_reason = "Pendiente (Sin comentario)"
                        open_pending += 1

            # Matchear metadatos de Tableau con Fallback a reporte_bonus.xlsx si es necesario.
            # Ambas fuentes de Tableau (WO Pending y Detail Data Implementation) sólo aplican
            # a WOs pendientes (FT Inprocessing): 'WO Pending' es un snapshot de lo que está
            # actualmente en proceso, así que no tiene sentido matchear WOs ya cerradas contra
            # esa vista. Las cerradas se enriquecen solo con datos propios de GNOC + NIMS.
            create_time_str = create_dt.strftime("%Y-%m-%d %H:%M:%S")
            tab_info = {} if is_closed else tableau_map.get(wo_code, {})

            if tab_info:
                matched_tableau_count += 1
                # Tableau a veces manda "FBB" (u otro valor genérico que no es ningún branch
                # real) en vez de dejarlo vacío -si se acepta tal cual, bloquea la cascada de
                # abajo (Caja -> NIMS/Site -> GNOC) porque branch ya "no está vacío". Se valida
                # contra KNOWN_BRANCHES para que un valor basura no gane prioridad sobre Caja/Site,
                # que sí resuelven el branch real (visto en WO_SPM_20260827_173274287: Tableau
                # decía "FBB", la caja LIC0313-... correspondía a LI4).
                branch_tableau = tab_info.get("branch", "")
                branch = branch_tableau if branch_tableau in KNOWN_BRANCHES else ""
                ticket_code = tab_info.get("ticket_code", "")
                wo_create_date = tab_info.get("wo_create_date", create_time_str)
                responsible_unit = tab_info.get("responsible_unit_wo", cd_group)
                account = account_gnoc or tab_info.get("account", "")
                if not account:
                    m_acc = re.search(r'\b(\d{2}_gftth_[a-zA-Z0-9_]+)\b', description, re.IGNORECASE)
                    if m_acc:
                        account = m_acc.group(1).lower()
                warranty_period = tab_info.get("warranty_period", "")
                implementation_test = tab_info.get("implementation_test", "")
                act_status = tab_info.get("act_status", "")
                sub_status = tab_info.get("sub_status", "")
                online_status = tab_info.get("online_status", "")
                ft_gnoc = tab_info.get("ft_gnoc", "")
                staf_team = tab_info.get("staf_team", "")
                connector_code = tab_info.get("connector_code", "")
                compcontent = tab_info.get("compcontent", "")
            else:
                # Fallback: Si no está en Tableau, buscamos datos en GNOC y reporte_bonus.xlsx
                ticket_code = str(row_vals[col_system_code]).strip() if row_vals[col_system_code] is not None else ""
                wo_create_date = create_time_str
                ft_gnoc = str(row_vals[col_ft]).strip() if row_vals[col_ft] is not None else ""
                # COMPCONTENT ("Descripción") es un dato propio de Tableau; si la WO no
                # matcheó ahí, se completa con la descripción propia de GNOC.
                compcontent = str(row_vals[col_description]).strip() if row_vals[col_description] is not None else ""

                # Cuenta de cliente: primero la columna "Subscribers" de GNOC, con
                # fallback a un regex sobre la descripción si esa columna viniera vacía.
                account = account_gnoc
                if not account:
                    m_acc = re.search(r'\b(\d{2}_gftth_[a-zA-Z0-9_]+)\b', description, re.IGNORECASE)
                    if m_acc:
                        account = m_acc.group(1).lower()
                    
                implementation_test = ""
                staf_team = ""
                warranty_period = ""
                branch = "" # Se derivará más abajo desde NIMS si está disponible
                responsible_unit = cd_group
                act_status = ""
                sub_status = ""
                online_status = ""
                connector_code = "" # Se derivará más abajo desde NIMS si está disponible

            # Fallback a reporte_bonus.xlsx (Detail Data Implementation) por Account cuando
            # falta el Implementation test y/o el Staf Team, sin importar si la WO matcheó
            # parcialmente en 'WO Pending' (por eso esto va fuera del if/else de arriba: antes
            # solo corría en la rama "no matcheó nada en WO Pending"). 'WO Pending' y 'Detail
            # Data Implementation' son vistas de Tableau distintas -una WO puede traer
            # implementation_test de una y necesitar el staf_team de la otra, así que ambos
            # campos se resuelven de forma independiente, no atados a la misma condición.
            # Igual que 'WO Pending', solo aplica a pendientes (FT Inprocessing), nunca a Close.
            if not is_closed and account and account in bonus_map and (not implementation_test or not staf_team):
                bonus_info = bonus_map[account]
                if not implementation_test:
                    implementation_test = bonus_info.get("implementation_test", "")
                if not staf_team:
                    staf_team = bonus_info.get("staf_team", "")
                impl_dt = parse_date(implementation_test)
                if impl_dt and create_dt:
                    warranty_period = str((create_dt - impl_dt).days)

            # Buscar en nims_subscribers para complementar datos si es necesario (ya sea porque no hay branch, connector_code o account)
            nims_box_db = ""
            nims_conn_db = ""
            nims_site_db = ""
            nims_unit_name_db = ""
            nims_account_db = ""
            
            if account or connector_code:
                # Reutilizar el cursor principal (cursor) para máxima velocidad
                cursor.execute("""
                    SELECT box_code, connector_code, site_code, unit_name, account 
                    FROM nims_subscribers 
                    WHERE (account = ? AND account != '') OR (connector_code = ? AND connector_code != '') 
                    LIMIT 1
                """, (account, connector_code))
                nims_row = cursor.fetchone()
                if nims_row:
                    nims_box_db, nims_conn_db, nims_site_db, nims_unit_name_db, nims_account_db = nims_row
                    
            # Completar account si falta y se encontró en NIMS
            if not account and nims_account_db:
                account = nims_account_db
            # "Connector / BOX": es la caja del cliente, no el splitter. NIMS manda primero
            # (el "Sub node code", nims_box_db) porque es la fuente más actualizada de en qué
            # caja está el cliente ahora mismo; si NIMS no tiene el dato, se usa el de Tableau.
            connector_code = nims_box_db or connector_code
                
            # Si la orden no tiene una cuenta de abonado (Account) asociada, se descarta por completo
            if not account or not str(account).strip():
                continue

            # Validar contra tm_subscribers en SQLite (caso insensitivo)
            has_tm = False
            tm_status = ""
            if account:
                # COLLATE NOCASE (en vez de LOWER(username) = LOWER(?)) para poder usar
                # idx_tm_username_nocase; LOWER() en ambos lados fuerza un full table scan
                # de tm_subscribers (68k filas) en cada WO, ~127x más lento por fila.
                cursor.execute("SELECT status FROM tm_subscribers WHERE username = ? COLLATE NOCASE", (account.strip(),))
                tm_row = cursor.fetchone()
                if tm_row:
                    has_tm = True
                    tm_status = str(tm_row[0]).strip()

            # Online Status: jerarquía de 3 niveles, siempre en este orden, sin importar
            # lo que haya traído Tableau (esta columna se resuelve 100% desde TMs/BRAS):
            #   1. No está en TMs                    -> CANCEL
            #   2. Está en TMs con status='2'         -> BLOCK BY DEBT (bloqueado por deuda)
            #   3. Está en TMs con cualquier otro status -> se deja vacío acá; se resuelve
            #      después de insertar en la BD con una consulta en vivo al BRAS
            #      (ver bras_status_check.py), que devuelve ONLINE / NOT ONLINE /
            #      NUNCA TUVO SERVICIO (esta última cuando el BRAS no tiene ningún
            #      registro de la cuenta).
            if not has_tm:
                online_status = "CANCEL"
                act_status = "CANCEL"
                sub_status = "CANCEL"
            elif tm_status == "2":
                online_status = "BLOCK BY DEBT"
            else:
                online_status = ""
                
            # Resolución de branch en cascada:
            #  1. Caja: la de Tableau (connector_code) si vino; si no, la de NIMS (Sub node code).
            #  2. Si falta el branch pero ya hay una caja (de cualquiera de las dos fuentes),
            #     se busca esa caja en 'boxes' (List of Boxes) para sacar su branch.
            #  3. Si tampoco hay caja en ningún lado, o la caja no está en List of Boxes,
            #     se cae al heurístico de prefijos existente (unit_name/site_code de NIMS).
            #  4. Último recurso, sin depender de NIMS: GNOC mismo manda el branch en
            #     responsible_unit/cd_group ("VTP_LI4BR FBB team") -cubre cuentas tan nuevas
            #     que NIMS aún no las tiene.
            box_code_for_branch = connector_code or nims_box_db or ""
            if not branch and box_code_for_branch:
                branch = boxes_branch_map.get(box_code_for_branch.strip().upper(), "")

            if not branch:
                ref_site = nims_site_db or connector_code or account or description or ""
                parsed_site_match = re.search(r'([A-Z]{2,4}\d{4,5})', ref_site)
                parsed_site_code = parsed_site_match.group(1) if parsed_site_match else ""
                branch = get_branch_from_nims(nims_unit_name_db, parsed_site_code)

            if not branch:
                # responsible_unit puede venir de Tableau (tab_info.get("responsible_unit_wo"))
                # con un texto genérico sin branch ("VTP - LIMA TGI FIXED BROADBAND/..."); cd_group
                # es siempre el dato crudo de GNOC en el formato que este parser espera
                # ("VTP_LI4BR FBB team"), así que se intenta también como alternativa.
                branch = get_branch_from_responsible_unit(responsible_unit) or get_branch_from_responsible_unit(cd_group)

            # Priorizar NIMS para la topología del cliente
            raw_ref = nims_conn_db or nims_box_db or nims_site_db or connector_code or account or description or ""
            topo = parse_box_code(raw_ref)

            processed_data.append((
                wo_code,
                wo_name,
                wo_type,
                description,
                wo_status,
                create_time_str,
                cd_group,
                ft_technician,
                priority,
                ft_comment,
                pending_hours,
                close_reason,
                is_error,
                branch,
                ticket_code,
                wo_create_date,
                responsible_unit,
                account,
                warranty_period,
                implementation_test,
                act_status,
                sub_status,
                online_status,
                ft_gnoc,
                staf_team,
                connector_code,
                compcontent,
                topo.get("site_code"),
                topo.get("xb_code"),
                topo.get("hb_code"),
                topo.get("line_code"),
                topo.get("box_code"),
                topo.get("expansion_code"),
                closed_time_str,
                resolution_hours
            ))

        # Guardar en base de datos
        print("Guardando registros consolidados en la base de datos SQLite...")
        cursor.execute("DELETE FROM work_orders")
        cursor.executemany("""
            INSERT INTO work_orders (
                wo_code, wo_name, wo_type, description, wo_status, 
                create_time, cd_group, ft_technician, priority, 
                ft_comment, pending_hours, close_reason, is_error,
                branch, ticket_code, wo_create_date, responsible_unit, account,
                warranty_period, implementation_test, act_status, sub_status,
                online_status, ft_gnoc, staf_team, connector_code, compcontent,
                site_code, xb_code, hb_code, line_code, box_code, expansion_code,
                closed_time, resolution_hours
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, processed_data)

        # Registrar fecha/hora de esta sincronización exitosa, para mostrarla en el dashboard
        # (ver /api/stats en server.py). CREATE TABLE IF NOT EXISTS por si esta base de datos
        # no pasó todavía por db_setup.py con esta tabla agregada.
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sync_meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        cursor.execute("""
            INSERT INTO sync_meta (key, value) VALUES ('last_sync_at', ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),))

        conn.commit()

        print("\n--- RESUMEN DE PROCESAMIENTO CONSOLIDADO ---")
        print(f"  - Total ordenes procesadas: {len(processed_data)}")
        print(f"  - Ordenes matcheadas con Tableau: {matched_tableau_count}")
        print(f"  - Errores de Marlo: {marlo_errors}")
        print(f"  - Comentarios clasificados con éxito: {classified_count}")
        print(f"  - Cerradas sin comentarios: {closed_without_comment}")
        print(f"  - Abiertas pendientes: {open_pending}")

    finally:
        conn.close()
        wb.close()
        for closed_wb in closed_wbs:
            closed_wb.close()

        for path in [temp_excel_path] + closed_temp_paths:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass

    # Nivel 3 de Online Status (ONLINE / NOT ONLINE / NUNCA TUVO SERVICIO): se resuelve
    # después de insertar, con una consulta en vivo al BRAS por cada cuenta pendiente.
    try:
        import bras_status_check
        bras_status_check.resolve_online_status_bulk()
    except Exception as e:
        print(f"[Aviso] No se pudo completar la verificación en vivo de Online Status (BRAS): {e}")

    print("\n¡Procesamiento y matcheo completo!")

if __name__ == "__main__":
    main()
