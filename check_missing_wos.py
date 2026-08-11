import openpyxl

sample_wos = [
    "WO_SPM_20260720_171331169",
    "WO_SPM_20260720_171331658",
    "WO_SPM_20260720_171332247",
    "WO_SPM_20260720_171332255",
    "WO_SPM_20260720_171332321",
    "WO_SPM_20260720_171332650"
]

print("--- BUSCANDO EN REPORTE GNOC ---")
wb_gnoc = openpyxl.load_workbook("reporte_gnoc.xlsx", data_only=True)
sheet_gnoc = wb_gnoc.active
gnoc_rows = list(sheet_gnoc.iter_rows(values_only=True))
header_gnoc = [str(val) for val in gnoc_rows[7]]
col_wo_gnoc = header_gnoc.index("WO code")
col_ft_gnoc = header_gnoc.index("FT")
col_st_gnoc = header_gnoc.index("WO Status")

for wo in sample_wos:
    found = False
    for r in gnoc_rows[8:]:
        if r and len(r) > col_wo_gnoc and str(r[col_wo_gnoc]).strip() == wo:
            print(f"GNOC -> WO: {wo}, Status: {r[col_st_gnoc]}, FT: {r[col_ft_gnoc]}")
            found = True
            break
    if not found:
        print(f"GNOC -> WO: {wo} NO ENCONTRADA EN GNOC")

print("\n--- BUSCANDO EN REPORTE TABLEAU ---")
wb_tab = openpyxl.load_workbook("reporte_tableau.xlsx", data_only=True)
sheet_tab = wb_tab.active
tab_rows = list(sheet_tab.iter_rows(values_only=True))
print(f"Total filas en reporte_tableau.xlsx: {len(tab_rows)}")

for wo in sample_wos:
    found = False
    for r in tab_rows[1:]:
        if r and len(r) > 2 and str(r[2]).strip() == wo:
            print(f"TABLEAU -> WO: {wo}, Branch: {r[1]}, Warranty: {r[8]}")
            found = True
            break
    if not found:
        print(f"TABLEAU -> WO: {wo} NO ENCONTRADA EN TABLEAU EXCEL")
