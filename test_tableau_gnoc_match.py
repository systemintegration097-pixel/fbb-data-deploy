import openpyxl

def test_match():
    print("Leyendo reporte_gnoc.xlsx...")
    wb_gnoc = openpyxl.load_workbook("reporte_gnoc.xlsx", data_only=True)
    sheet_gnoc = wb_gnoc.active
    header_gnoc = [cell.value for cell in sheet_gnoc[8]]
    col_gnoc_wo = header_gnoc.index("WO code") if "WO code" in header_gnoc else 1
    col_gnoc_status = header_gnoc.index("WO Status") if "WO Status" in header_gnoc else 6
    col_gnoc_ft = header_gnoc.index("FT") if "FT" in header_gnoc else 13
    
    gnoc_data = {}
    for row in sheet_gnoc.iter_rows(min_row=9, values_only=True):
        if row and len(row) > col_gnoc_wo and row[col_gnoc_wo]:
            wo = str(row[col_gnoc_wo]).strip()
            status = str(row[col_gnoc_status]).strip() if len(row) > col_gnoc_status and row[col_gnoc_status] else ""
            ft = str(row[col_gnoc_ft]).strip() if len(row) > col_gnoc_ft and row[col_gnoc_ft] else ""
            gnoc_data[wo] = {"status": status, "ft": ft}
            
    print(f"Total WOs en GNOC: {len(gnoc_data)}")

    print("\nLeyendo reporte_tableau.xlsx...")
    wb_tab = openpyxl.load_workbook("reporte_tableau.xlsx", data_only=True)
    sheet_tab = wb_tab.active
    rows_tab = list(sheet_tab.iter_rows(values_only=True))
    header_tab = rows_tab[0] if len(rows_tab) > 0 else []
    print("Header Tableau:", header_tab)
    
    tab_data = {}
    for r in rows_tab[1:]:
        if r and len(r) > 2 and r[2]: # WO_CODE es la col 2 (0-indexed)
            wo = str(r[2]).strip()
            branch = str(r[1]).strip() if r[1] is not None else ""
            warranty = str(r[8]).strip() if len(r) > 8 and r[8] is not None else ""
            impl_test = str(r[9]).strip() if len(r) > 9 and r[9] is not None else ""
            act_status = str(r[10]).strip() if len(r) > 10 and r[10] is not None else ""
            sub_status = str(r[11]).strip() if len(r) > 11 and r[11] is not None else ""
            connector = str(r[15]).strip() if len(r) > 15 and r[15] is not None else ""
            tab_data[wo] = {
                "branch": branch,
                "warranty": warranty,
                "impl_test": impl_test,
                "act_status": act_status,
                "sub_status": sub_status,
                "connector": connector
            }

    print(f"Total WOs en Tableau: {len(tab_data)}")

    # Matchear
    matched_in_gnoc = 0
    closed_in_gnoc = 0
    inprocessing_in_gnoc = 0
    other_status_gnoc = 0
    not_in_gnoc = 0

    for wo, tab_info in tab_data.items():
        if wo in gnoc_data:
            matched_in_gnoc += 1
            st = gnoc_data[wo]["status"].lower()
            if "close" in st:
                closed_in_gnoc += 1
            elif "inprocessing" in st or "processing" in st or "wait" in st:
                inprocessing_in_gnoc += 1
            else:
                other_status_gnoc += 1
        else:
            not_in_gnoc += 1

    print("\n--- RESULTADO DE CRUCE GNOC vs TABLEAU ---")
    print(f"WOs de Tableau encontradas en GNOC: {matched_in_gnoc}")
    print(f"  -> En GNOC figuran realmente CERRADAS (Close/Closed): {closed_in_gnoc}")
    print(f"  -> En GNOC figuran PENDIENTES (Inprocessing / Processing / Wait): {inprocessing_in_gnoc}")
    print(f"  -> En GNOC figuran en OTRO ESTADO: {other_status_gnoc}")
    print(f"WOs de Tableau NO encontradas en GNOC (Nuevas/Fuera de rango): {not_in_gnoc}")

if __name__ == '__main__':
    test_match()
