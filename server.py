import os
import glob
import sqlite3
import io
import sys
import time
import json
from datetime import datetime
import openpyxl
import urllib.request
import urllib.error
from flask import Flask, jsonify, request, send_from_directory, send_file
import process_data
import nims_topology
import kpi_calc
import daily_report
import deploy_pending
import cloud_sync
import sheets_push
try:
    import an_portal_client
except Exception:
    an_portal_client = None
import olt_audit

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "DataBaseFBB"))
from db_manager import DBManager as FBBManager
import db_importer as fbb_importer


app = Flask(__name__, static_folder="static")
DB_PATH = "gnoc.db"
OLT_INPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "olts_input.xlsx")

def get_db_connection():
    # timeout=30: gnoc.db tiene varios escritores (este server, download_nims.py,
    # process_data.py); el default de sqlite3 (5s) hace que cualquier escritura que se
    # cruce con otra falle con "database is locked" en vez de simplemente esperar.
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn

# Servir index principal
@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")

import sys
import subprocess
import threading
import concurrent.futures

# Cuando stdout no está atado a una consola interactiva real (redirigido a archivo, pipe,
# lanzado como servicio, etc.), Python en Windows cae al codepage legado ('charmap', ej.
# cp1252) en vez de UTF-8. Los scripts de sync (download_report.py, etc.) imprimen texto en
# español con tildes/ñ/¡/¿; si esos bytes no decodifican bien como UTF-8 al leerlos (ver
# run_sync_script) aparece el caracter de reemplazo U+FFFD, y al intentar volver a imprimir
# ESE caracter aquí con 'charmap' Python lanza UnicodeEncodeError y rompe todo el sync.
# errors='replace' hace que cualquier caracter no representable se sustituya en vez de crashear.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

sync_status = {
    "state": "idle", # "idle", "downloading", "processing", "success", "error"
    "message": ""
}

sync_lock = threading.Lock()
active_sync_processes = []

def run_sync_script(script_name, task_name, env_overrides=None):
    script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)
    env = os.environ.copy()
    # Forzar modo UTF-8 de Python en el subproceso: sin esto, en Windows el hijo puede
    # escribir su stdout en el codepage legado de la consola (cp1252) en vez de UTF-8,
    # que es lo que este proceso asume al leerlo abajo (encoding="utf-8") -> caracteres
    # como tildes/ñ/¡/¿ salen corruptos (U+FFFD) y pueden hacer crashear el print() del padre.
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    if env_overrides:
        env.update(env_overrides)
    # Ejecutar en proceso independiente para aislamiento completo de Playwright y asyncio
    proc = subprocess.Popen(
        [sys.executable, "-u", script_path],
        cwd=os.path.dirname(os.path.abspath(__file__)),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
        env=env
    )
    with sync_lock:
        active_sync_processes.append(proc)
    
    output_lines = []
    try:
        for line in iter(proc.stdout.readline, ''):
            line_str = line.strip()
            if line_str:
                print(f"[{task_name}] {line_str}", flush=True)
                output_lines.append(line_str)
    finally:
        proc.stdout.close()
        proc.wait()
        with sync_lock:
            if proc in active_sync_processes:
                active_sync_processes.remove(proc)

    if proc.returncode != 0:
        err_snippet = "\n".join(output_lines[-4:]) if output_lines else f"Código de salida {proc.returncode}"
        raise Exception(f"{task_name}: {err_snippet}")
    return True

def compute_filter_create_time_for_months(from_month, to_month):
    """from_month/to_month en formato 'YYYY-MM'. Devuelve el string de rango que espera
    download_report.py ('DD/MM/YYYY HH:MM:SS to DD/MM/YYYY HH:MM:SS'), cubriendo desde el
    primer día de from_month 00:00:00 hasta el último día de to_month 23:59:59."""
    import calendar
    start_dt = datetime.strptime(from_month, "%Y-%m")
    end_first = datetime.strptime(to_month, "%Y-%m")
    last_day = calendar.monthrange(end_first.year, end_first.month)[1]
    end_dt = end_first.replace(day=last_day, hour=23, minute=59, second=59)
    return f"{start_dt.strftime('%d/%m/%Y %H:%M:%S')} to {end_dt.strftime('%d/%m/%Y %H:%M:%S')}"

# El rango de meses elegido en el dashboard solo vivía en el body del POST /api/sync (una
# sola corrida) y en localStorage del navegador (solo cosmético, para que el selector no se
# vea vacío) -- el ciclo automático (_auto_excel_sync_loop) nunca se enteraba y siempre volvía
# a usar el FILTER_CREATE_TIME estático de .env. Este archivo persiste la última elección para
# que tanto una sincronización manual como el ciclo automático usen el mismo rango hasta que
# alguien lo cambie explícitamente.
SYNC_MONTH_RANGE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sync_month_range.json")

def save_sync_month_range(from_month, to_month):
    with open(SYNC_MONTH_RANGE_PATH, "w", encoding="utf-8") as f:
        json.dump({"from_month": from_month, "to_month": to_month}, f)

def load_sync_month_range():
    if not os.path.exists(SYNC_MONTH_RANGE_PATH):
        return None, None
    try:
        with open(SYNC_MONTH_RANGE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("from_month") or "", data.get("to_month") or ""
    except (ValueError, OSError):
        return None, None

def clear_sync_month_range():
    if os.path.exists(SYNC_MONTH_RANGE_PATH):
        os.remove(SYNC_MONTH_RANGE_PATH)

def run_background_sync(gnoc_env_overrides=None):
    global sync_status
    try:
        # Eliminar archivos Excel antiguos para evitar usar datos obsoletos si falla la descarga
        old_files = ["reporte_tableau.xlsx", "reporte_gnoc.xlsx", "reporte_bonus.xlsx", "reporte_cnoc.xlsx"]
        old_files.extend(glob.glob("reporte_gnoc_closed_*.xlsx"))
        for old_file in old_files:
            if os.path.exists(old_file):
                try:
                    os.remove(old_file)
                    print(f"Archivo obsoleto {old_file} eliminado para asegurar datos frescos.")
                except Exception as e:
                    print(f"No se pudo eliminar {old_file}: {e}")
        
        with sync_lock:
            sync_status["state"] = "downloading"
            range_note = ""
            if gnoc_env_overrides and gnoc_env_overrides.get("FILTER_CREATE_TIME"):
                range_note = f" (rango GNOC: {gnoc_env_overrides['FILTER_CREATE_TIME']})"
            sync_status["message"] = f"Descargando GNOC, Tableau y NIMS en paralelo...{range_note}"

        # GNOC y Tableau son sesiones independientes en portales distintos ejecutadas en
        # subprocesos aislados para evitar conflictos con el event loop de Playwright/asyncio.
        # CNOC recibe el mismo rango de fechas que GNOC (gnoc_env_overrides), pero su fallo NO
        # es fatal para el resto del sync: es una fuente complementaria (WOs que no aparecen en
        # GNOC), más nueva y con un login más inestable, así que si falla simplemente se sigue
        # con los datos de CNOC que ya existan en disco de una corrida anterior (o ninguno).
        tasks = {
            "GNOC": ("download_report.py", gnoc_env_overrides),
            "Tableau": ("download_tableau.py", None),
            "NIMS": ("download_nims.py", None),
            "CNOC": ("download_cnoc.py", gnoc_env_overrides),
        }
        errors = []
        cnoc_error = None
        with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
            future_to_name = {
                executor.submit(run_sync_script, script_file, name, env_overrides): name
                for name, (script_file, env_overrides) in tasks.items()
            }
            for future in concurrent.futures.as_completed(future_to_name):
                name = future_to_name[future]
                try:
                    future.result()
                    print(f"[Sync] Descarga de {name} finalizada con éxito.", flush=True)
                except Exception as e:
                    print(f"[Sync] Descarga de {name} falló: {e}", flush=True)
                    if name == "CNOC":
                        cnoc_error = str(e)
                    else:
                        errors.append(str(e))

        if errors:
            raise Exception("Fallaron una o más descargas -> " + " | ".join(errors))

        # Paso 4: Consolidar y procesar datos en SQLite (en subproceso para aislamiento de memoria y CPU)
        with sync_lock:
            sync_status["state"] = "processing"
            sync_status["message"] = "Procesando y cruzando datos GNOC, Tableau, NIMS y CNOC..."

        run_sync_script("process_data.py", "ProcessData")
        kpi_calc.invalidate_cache()

        with sync_lock:
            sync_status["state"] = "success"
            if cnoc_error:
                sync_status["message"] = (
                    "Sincronización completada con éxito (CNOC falló y se omitió esta vez: "
                    + cnoc_error[:200] + ")"
                )
            else:
                sync_status["message"] = "Sincronización completada con éxito."
    except Exception as e:
        with sync_lock:
            sync_status["state"] = "error"
            sync_status["message"] = f"Error en la sincronización: {str(e)}"

# Endpoint: Consultar Estado de Usuario en BRAS
@app.route("/api/bras/user_info", methods=["GET"])
def get_bras_user_info():
    account = request.args.get("account", "").strip()
    bras = request.args.get("bras", "").strip()
    if not account or not bras:
        return jsonify({"error": "Debe especificar los parámetros 'account' y 'bras'."}), 400
        
    url = f"http://10.121.62.102:8080/backup/cgi-bin/bras_checkuser/bras_checkkick_online.php?cat=view&acc={account}&domain=&bras={bras}"
    try:
        req = urllib.request.Request(
            url,
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=12) as response:
            text = response.read().decode('utf-8', errors='ignore')
            
            # Analizar y parsear el estado Online del suscriptor
            try:
                resolved_status = "NOT ONLINE"
                text_upper = text.upper()
                
                is_online = False
                for line in text_upper.splitlines():
                    if "ONLINE" in line and "NOT ONLINE" not in line:
                        is_online = True
                        break
                        
                if is_online:
                    import re
                    ip_match = re.search(r'IPV4-ADDRESS\s*:\s*([0-9.]+)', text_upper)
                    if ip_match:
                        ip = ip_match.group(1).strip()
                        if ip.startswith("10."):
                            resolved_status = "ONLINE"
                        else:
                            resolved_status = "ONLINE IP NOK"
                    else:
                        resolved_status = "ONLINE"
                
                # Actualizar el estado en la tabla de órdenes de trabajo de SQLite
                db_conn = get_db_connection()
                try:
                    db_conn.execute("""
                        UPDATE work_orders
                        SET online_status = ?
                        WHERE account = ?
                    """, (resolved_status, account))
                    db_conn.commit()
                    print(f"[BRAS Sync] Estado online_status de '{account}' actualizado a '{resolved_status}'.", flush=True)
                finally:
                    db_conn.close()
            except Exception as db_err:
                print(f"[BRAS Sync Error] Error al actualizar base de datos: {db_err}", flush=True)
                
            return jsonify({"success": True, "output": text})
    except Exception as e:
        # Si la URL no existe, hay un error de conexión, o cualquier otro fallo, se asume que el cliente está CANCEL
        try:
            db_conn = get_db_connection()
            try:
                db_conn.execute("""
                    UPDATE work_orders
                    SET online_status = 'CANCEL'
                    WHERE account = ?
                """, (account,))
                db_conn.commit()
                print(f"[BRAS Sync] Estado de '{account}' actualizado a 'CANCEL' debido a fallo de conexión o URL no existente.", flush=True)
            finally:
                db_conn.close()
        except Exception as db_err:
            print(f"[BRAS Sync Error] Error al guardar estado CANCEL: {db_err}", flush=True)
            
        return jsonify({"success": True, "output": f"NOT ONLINE (CANCEL - Error: {str(e)})"})

# Endpoint: Consultar Estado de ONU en OLT (AN-Portal)
@app.route("/api/an-portal/onu_status", methods=["GET"])
def get_onu_status():
    mac = request.args.get("mac", "").strip()
    olt = request.args.get("olt", "").strip()
    if not mac or not olt:
        return jsonify({"error": "Debe especificar los parámetros 'mac' y 'olt'."}), 400
        
    res = an_portal_client.query_onu_status_from_portal(mac, olt)
    return jsonify(res)

# Endpoint: Consultar Información del Suscriptor (NIMS + TMs)
@app.route("/api/subscriber/info", methods=["GET"])
def get_subscriber_info():
    account = request.args.get("account", "").strip()
    if not account:
        return jsonify({"error": "Debe especificar el parámetro 'account'."}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Consultar en nims_subscribers
        cursor.execute("""
            SELECT account, customer_name, phone, status, site_code, site_name, box_code, connector_code, address, connection_date
            FROM nims_subscribers
            WHERE account = ? OR account = ?
        """, (account, "04_gftth_" + account if not account.startswith("04_gftth_") else account))
        nims_row = cursor.fetchone()
        
        # Consultar en tm_subscribers
        cursor.execute("""
            SELECT username, mac, port, status, activedate, canceldate, suspenddate, reactivedate, bras, ipaddress
            FROM tm_subscribers
            WHERE username = ? OR username = ?
        """, (account, "04_gftth_" + account if not account.startswith("04_gftth_") else account))
        tm_row = cursor.fetchone()
        
        nims_data = None
        if nims_row:
            nims_data = {
                "account": nims_row["account"],
                "customer_name": nims_row["customer_name"],
                "phone": nims_row["phone"],
                "status": nims_row["status"],
                "site_code": nims_row["site_code"],
                "site_name": nims_row["site_name"],
                "box_code": nims_row["box_code"],
                "connector_code": nims_row["connector_code"],
                "address": nims_row["address"],
                "connection_date": nims_row["connection_date"]
            }
            
        tm_data = None
        if tm_row:
            tm_data = {
                "username": tm_row["username"],
                "mac": tm_row["mac"],
                "port": tm_row["port"],
                "status": tm_row["status"],
                "activedate": tm_row["activedate"],
                "canceldate": tm_row["canceldate"],
                "suspenddate": tm_row["suspenddate"],
                "reactivedate": tm_row["reactivedate"],
                "bras": tm_row["bras"],
                "ipaddress": tm_row["ipaddress"]
            }
            
        return jsonify({
            "success": True,
            "nims": nims_data,
            "tms": tm_data
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# Endpoint: Consultar Topología de Red NIMS
@app.route("/api/nims/topology", methods=["GET"])
def get_nims_topology():
    query = request.args.get("query", "").strip()
    if not query:
        return jsonify({"error": "Debe especificar un término de consulta (ej: query=ARE0001 o query=SB111)"}), 400
        
    parsed = nims_topology.parse_box_code(query)
    
    # Si es un código de Sitio (ej: ARE0001), incluir el árbol completo del Sitio
    site_tree = None
    if parsed and parsed.get("site_code") and (query.upper() == parsed["site_code"] or not parsed.get("box_code")):
        site_tree = nims_topology.generate_site_topology(parsed["site_code"])
        
    # Obtener suscriptores agrupados por caja en este Sitio/OLT
    subscribers_by_box = {}
    if parsed and parsed.get("site_code"):
        site_code = parsed["site_code"]
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT n.account, n.box_code, n.customer_name, n.status as nims_status, 
                       t.mac, t.bras, t.status as tm_status
                FROM nims_subscribers n
                LEFT JOIN tm_subscribers t ON t.username = n.account
                WHERE n.site_code = ? OR n.site_code = ? OR n.site_code = ?
            """, (site_code, site_code + "OLT01", site_code.replace("OLT01", "")))
            
            for row in cursor.fetchall():
                box = row["box_code"]
                if not box:
                    continue
                box_clean = box.strip().upper()
                box_parts = box_clean.split("-")
                box_name = box_parts[-1] if len(box_parts) > 1 else box_clean
                
                # Limpiar si tiene terminación SPxx
                if "SP" in box_name:
                    box_name = box_name.split("SP")[0]
                    
                if box_name not in subscribers_by_box:
                    subscribers_by_box[box_name] = []
                    
                subscribers_by_box[box_name].append({
                    "account": row["account"],
                    "customer_name": row["customer_name"] or "Cliente Desconocido",
                    "nims_status": row["nims_status"] or "Active",
                    "mac": row["mac"] or "",
                    "bras": row["bras"] or "",
                    "tm_status": row["tm_status"] or ""
                })
        except Exception as e:
            print(f"Error querying subscribers for topology: {e}")
        finally:
            conn.close()
            
    return jsonify({
        "parsed": parsed,
        "site_tree": site_tree,
        "subscribers_by_box": subscribers_by_box
    })

# Endpoint: Reporte de Averías por Caja y Sitio
@app.route("/api/nims/faults_report", methods=["GET"])
def get_nims_faults_report():
    try:
        report = nims_topology.analyze_faults_by_topology(DB_PATH)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Endpoint: Reporte SLA de Pendientes por Branch
@app.route("/api/reports/branch_sla", methods=["GET"])
def get_branch_sla_report():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT 
                COALESCE(NULLIF(branch, ''), 'SIN BRANCH') as branch_name,
                SUM(CASE WHEN pending_hours > 0 AND pending_hours <= 24 THEN 1 ELSE 0 END) as under_24h,
                SUM(CASE WHEN pending_hours > 24 AND pending_hours <= 48 THEN 1 ELSE 0 END) as under_48h,
                SUM(CASE WHEN pending_hours > 48 AND pending_hours <= 72 THEN 1 ELSE 0 END) as under_72h,
                SUM(CASE WHEN pending_hours > 72 THEN 1 ELSE 0 END) as over_72h,
                COUNT(*) as total_pending
            FROM work_orders
            WHERE wo_status NOT IN ('Close', 'Closed', 'Closed FT', 'FT completed') AND is_error = 0
            GROUP BY branch_name
            ORDER BY total_pending DESC
        """)
        rows = cursor.fetchall()
        report = []
        for r in rows:
            report.append({
                "branch": r["branch_name"],
                "under_24h": r["under_24h"],
                "under_48h": r["under_48h"],
                "under_72h": r["under_72h"],
                "over_72h": r["over_72h"],
                "total_pending": r["total_pending"]
            })
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

# Endpoint: Reporte de Tipificación (Motivos de Cierre)
# Filtro opcional ?months=2026-06,2026-07 (por mes de creación de la WO, formato YYYY-MM)
@app.route("/api/reports/typification", methods=["GET"])
def get_typification_report():
    months_param = request.args.get("months", "").strip()
    months = [m.strip() for m in months_param.split(",") if m.strip()]

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Agrupado también por mes (además de por motivo) para poder armar columnas por
        # mes en la tabla del dashboard, sin perder el total/pendientes agregado de siempre.
        query = """
            SELECT
                COALESCE(NULLIF(close_reason, ''), 'PENDIENTE') as reason_name,
                strftime('%Y-%m', create_time) as month_key,
                COUNT(*) as total_wos,
                SUM(CASE WHEN wo_status NOT IN ('Close', 'Closed', 'Closed FT', 'FT completed') THEN 1 ELSE 0 END) as pending_wos
            FROM work_orders
        """
        params = []
        if months:
            placeholders = ",".join(["?"] * len(months))
            query += f" WHERE strftime('%Y-%m', create_time) IN ({placeholders})"
            params.extend(months)
        query += " GROUP BY reason_name, month_key"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        by_reason = {}
        for r in rows:
            reason = r["reason_name"]
            entry = by_reason.setdefault(reason, {"reason": reason, "total_wos": 0, "pending_wos": 0, "by_month": {}})
            entry["total_wos"] += r["total_wos"]
            entry["pending_wos"] += r["pending_wos"]
            if r["month_key"]:
                entry["by_month"][r["month_key"]] = entry["by_month"].get(r["month_key"], 0) + r["total_wos"]

        report = sorted(by_reason.values(), key=lambda x: -x["total_wos"])
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

# Endpoint: Diagnóstico de Caja de Distribución (Telecom Engineer)
@app.route("/api/nims/box_diagnostics", methods=["GET"])
def get_box_diagnostics():
    site_code = request.args.get("site_code", "").strip()
    box_code = request.args.get("box_code", "").strip()
    if not site_code or not box_code:
        return jsonify({"error": "Debe especificar los parámetros 'site_code' y 'box_code'."}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # 1. Obtener la ruta jerárquica teórica
        topo = nims_topology.parse_box_code(f"{site_code}_{box_code}")
        full_route = topo.get("full_route") if topo else f"{site_code} -> {box_code}"
        
        # 2. Buscar todos los suscriptores de esta caja en NIMS y cruzarlos con TMs
        cursor.execute("""
            SELECT n.account, n.customer_name, n.phone, n.status as nims_status, n.connector_code,
                   t.mac, t.bras, t.port as olt_port, t.status as tm_status
            FROM nims_subscribers n
            LEFT JOIN tm_subscribers t ON t.username = n.account
            WHERE (n.site_code = ? OR n.site_code = ? OR n.site_code = ?)
              AND n.box_code LIKE '%' || ? || '%'
        """, (site_code, site_code + "OLT01", site_code.replace("OLT01", ""), box_code))
        
        subs_rows = cursor.fetchall()
        clients = []
        for r in subs_rows:
            # Extraer número de puerto asignado del Splitter (por ejemplo, del conector code SP03 -> puerto 3 del splitter)
            conn_code = r["connector_code"] or ""
            splitter_port = "N/A"
            if "SP" in conn_code:
                parts = conn_code.split("SP")
                if len(parts) > 1 and parts[1].isdigit():
                    splitter_port = str(int(parts[1]))
            
            clients.append({
                "account": r["account"],
                "customer_name": r["customer_name"] or "Cliente Desconocido",
                "phone": r["phone"] or "N/A",
                "nims_status": r["nims_status"] or "Active",
                "mac": r["mac"] or "",
                "bras": r["bras"] or "",
                "olt_port": r["olt_port"] or "N/A",
                "splitter_port": splitter_port
            })
            
        # 3. Buscar todas las órdenes de trabajo activas/pendientes en esta caja
        cursor.execute("""
            SELECT wo_code, wo_name, wo_type, priority, create_time, pending_hours, ft_technician, ft_comment, close_reason
            FROM work_orders
            WHERE (site_code = ? OR site_code = ? OR site_code = ?)
              AND (box_code LIKE '%' || ? || '%' OR connector_code LIKE '%' || ? || '%')
              AND wo_status NOT IN ('Close', 'Closed', 'Closed FT', 'FT completed')
        """, (site_code, site_code + "OLT01", site_code.replace("OLT01", ""), box_code, box_code))
        
        wos_rows = cursor.fetchall()
        wos = []
        for r in wos_rows:
            wos.append({
                "wo_code": r["wo_code"],
                "wo_name": r["wo_name"],
                "wo_type": r["wo_type"],
                "priority": r["priority"],
                "create_time": r["create_time"],
                "pending_hours": round(r["pending_hours"], 1),
                "ft_technician": r["ft_technician"] or "Sin asignar",
                "ft_comment": r["ft_comment"] or "Sin comentarios",
                "classification": r["close_reason"] or "Pendiente"
            })
            
        # 4. Lógica de Diagnóstico de Ingeniería de Telecomunicaciones
        total_clients = len(clients)
        active_wos = len(wos)
        
        diagnosis = "NORMAL - Caja de distribución sin averías masivas reportadas."
        proposed_solution = "Monitoreo preventivo de alarmas ópticas en el sistema EMS. No se requiere intervención inmediata en campo."
        severity = "info"
        
        if active_wos >= 3 or (total_clients > 0 and active_wos / total_clients >= 0.5):
            severity = "critical"
            diagnosis = f"CRÍTICO - Alta probabilidad de corte de fibra multipar o daño severo en el Splitter de la caja {box_code}. Afectación del {round((active_wos/total_clients)*100 if total_clients > 0 else 100)}% de los abonados."
            proposed_solution = f"Desplegar cuadrilla de planta externa urgente. Realizar medición reflectométrica (OTDR) desde la OLT {site_code} hacia la ruta del distribuidor secundario para localizar corte o atenuación puntual. Inspeccionar físicamente la caja {box_code} y re-empalmar conectores en splitter."
        elif active_wos >= 1:
            severity = "warning"
            diagnosis = f"ALERTA - Avería de última milla detectada. Incidente focalizado que afecta a {active_wos} cliente(s) en la caja {box_code}."
            proposed_solution = f"Programar visita técnica domiciliaria para revisión de la acometida óptica (drop cable) de los clientes afectados. Limpiar y verificar conector mecánico en splitter de caja {box_code} y medir potencia óptica de recepción en ONT (rango aceptable: -15 a -25 dBm)."
            
        return jsonify({
            "success": True,
            "box_code": box_code,
            "site_code": site_code,
            "full_route": full_route,
            "total_clients": total_clients,
            "active_wos_count": active_wos,
            "clients": clients,
            "wos": wos,
            "severity": severity,
            "diagnosis": diagnosis,
            "proposed_solution": proposed_solution
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

# Endpoint: Inspector de Topología - Escaneo en vivo GPON + Estado BRAS
# Para un SITE + OLT (+ HUBBOX opcional) dado, arma la grilla teórica de cajas
# (HB01-04 x Línea 1-4 x SB/EB + EX) igual que nims_topology.get_boxes_for_line,
# busca los abonados ACTIVOS de NIMS en cada caja y valida en vivo contra BRAS
# cuántos están ONLINE/NOT ONLINE. Una caja "caída" (alarma) es una caja con
# clientes activos donde la mayoría está NOT ONLINE (%online <= 20).
BRAS_CHECK_URL = "http://10.121.62.102:8080/backup/cgi-bin/bras_checkuser/bras_checkkick_online.php"

def _check_bras_status_once(account, bras):
    """Un solo intento de consulta a BRAS. Devuelve (estado, ip_status) donde estado es
    ONLINE / NOT ONLINE / UNKNOWN (respuesta ambigua o error de red, NO es lo mismo que
    NOT ONLINE, así que se distingue para no generar alarmas falsas por un fallo puntual)
    / NUNCA ONLINE (la cuenta no tiene BRAS asignado en TMs: nunca tuvo sesión, no es
    una caída sino un cliente que nunca llegó a tener servicio - confirmado con el
    usuario 2026-07-27 para el caso real '04_gftth_gabrielak9')."""
    import re
    if not bras:
        return "NUNCA ONLINE", "IP NOK"
    url = f"{BRAS_CHECK_URL}?cat=view&acc={account}&domain=&bras={bras}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            text = response.read().decode('utf-8', errors='ignore')
            text_upper = text.upper()

            # Igual que /api/bras/user_info: se busca línea por línea (no un substring
            # sobre todo el texto) porque la página puede traer texto de ayuda/leyenda
            # que menciona "NOT ONLINE" en otra sección aunque la cuenta SÍ esté online.
            estado = "UNKNOWN"
            for line in text_upper.splitlines():
                if "NOT ONLINE" in line or "NO SESSION" in line:
                    estado = "NOT ONLINE"
                    break
                if "ONLINE" in line:
                    estado = "ONLINE"
                    break

            ip_match = re.search(r'IPV4[-_\s]*ADDRESS\s*:\s*([\d.]+)', text_upper)
            if ip_match and not ip_match.group(1).startswith(("172.", "9.", "0.")):
                ip_status = "IP OK"
            else:
                ip_status = "IP NOK"

            return estado, ip_status
    except Exception:
        return "UNKNOWN", "IP NOK"

def _check_bras_status(account, bras):
    """Consulta en vivo el estado de una cuenta en BRAS, con reintento si el primer intento
    no da ONLINE. Un solo timeout/blip de red no debería bastar para marcar una caja como
    caída — por eso se reintenta antes de aceptar un NOT ONLINE/UNKNOWN como definitivo.
    NUNCA ONLINE no se reintenta: es un hecho de los datos (sin BRAS asignado en TMs),
    no una condición transitoria que un reintento pueda cambiar."""
    import time
    estado, ip_status = _check_bras_status_once(account, bras)
    if estado not in ("ONLINE", "NUNCA ONLINE"):
        time.sleep(1)
        estado, ip_status = _check_bras_status_once(account, bras)
    return account, estado, ip_status

def _color_for_pct(pct):
    if pct <= 20:
        return "red"
    elif pct < 80:
        return "orange"
    return "green"

# El Inspector Topología saca los clientes activos + BRAS directo de "Port Release
# Guide.xlsx" (hojas NIMS y TMs, ya cruzadas ahí por el usuario) en vez de gnoc.db:
# es una fuente más fresca/confiable para esta función específica. El archivo pesa
# ~40MB, así que se cachea en un sqlite local y solo se re-importa cuando cambia su
# fecha de modificación (mismo patrón que gpon5.py: sync_excel_to_db + mod_time).
PORT_RELEASE_GUIDE_PATH = r"C:\Users\jjvar\OneDrive\Documentos\Port Release Guide.xlsx"
PRG_CACHE_DB = "port_release_cache.db"

def _prg_file_mod_time():
    return os.path.getmtime(PORT_RELEASE_GUIDE_PATH) if os.path.exists(PORT_RELEASE_GUIDE_PATH) else 0

def _prg_cached_mod_time():
    if not os.path.exists(PRG_CACHE_DB):
        return 0
    try:
        conn = sqlite3.connect(PRG_CACHE_DB)
        cur = conn.cursor()
        cur.execute("SELECT value FROM metadata WHERE key = 'mod_time'")
        row = cur.fetchone()
        conn.close()
        return float(row[0]) if row else 0
    except Exception:
        return 0

def _prg_sync_cache():
    wb = openpyxl.load_workbook(PORT_RELEASE_GUIDE_PATH, read_only=True, data_only=True)
    try:
        nims_sheet = wb["NIMS"]
        tms_sheet = wb["TMs"]

        conn = sqlite3.connect(PRG_CACHE_DB)
        cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS prg_nims")
        cur.execute("CREATE TABLE prg_nims (account TEXT, box_code TEXT, department TEXT, customer_name TEXT, phone TEXT)")
        cur.execute("DROP TABLE IF EXISTS prg_tms")
        cur.execute("CREATE TABLE prg_tms (username TEXT PRIMARY KEY, bras TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT)")

        nims_header = next(nims_sheet.iter_rows(max_row=1, values_only=True))
        idx = {str(name).strip(): i for i, name in enumerate(nims_header) if name}
        col_status, col_account = idx.get("Account Status"), idx.get("Account")
        col_name, col_phone = idx.get("Full name"), idx.get("Tel")
        col_box, col_dept = idx.get("Sub node code"), idx.get("Department")

        batch = []
        for row in nims_sheet.iter_rows(min_row=2, values_only=True):
            status = row[col_status] if col_status is not None and len(row) > col_status else None
            if not status or str(status).strip().lower() != "activo":
                continue
            account = row[col_account] if col_account is not None and len(row) > col_account else None
            box = row[col_box] if col_box is not None and len(row) > col_box else None
            if not account or not box:
                continue
            dept = row[col_dept] if col_dept is not None and len(row) > col_dept else ""
            name = row[col_name] if col_name is not None and len(row) > col_name else ""
            phone = row[col_phone] if col_phone is not None and len(row) > col_phone else ""
            batch.append((
                str(account).strip().lower(), str(box).strip().upper(), str(dept).strip() if dept else "",
                str(name).strip() if name else "", str(phone).strip() if phone else ""
            ))
            if len(batch) >= 5000:
                cur.executemany("INSERT INTO prg_nims VALUES (?,?,?,?,?)", batch)
                batch = []
        if batch:
            cur.executemany("INSERT INTO prg_nims VALUES (?,?,?,?,?)", batch)

        tms_header = next(tms_sheet.iter_rows(max_row=1, values_only=True))
        idx_t = {str(name).strip(): i for i, name in enumerate(tms_header) if name}
        col_user, col_bras = idx_t.get("USERNAME"), idx_t.get("BRAS")

        batch = []
        for row in tms_sheet.iter_rows(min_row=2, values_only=True):
            user = row[col_user] if col_user is not None and len(row) > col_user else None
            if not user:
                continue
            bras = row[col_bras] if col_bras is not None and len(row) > col_bras else ""
            batch.append((str(user).strip().lower(), str(bras).strip() if bras else ""))
            if len(batch) >= 5000:
                cur.executemany("INSERT OR REPLACE INTO prg_tms VALUES (?,?)", batch)
                batch = []
        if batch:
            cur.executemany("INSERT OR REPLACE INTO prg_tms VALUES (?,?)", batch)

        cur.execute("CREATE INDEX IF NOT EXISTS idx_prg_nims_box ON prg_nims(box_code)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_prg_nims_account ON prg_nims(account)")
        cur.execute("INSERT OR REPLACE INTO metadata (key, value) VALUES ('mod_time', ?)", (str(_prg_file_mod_time()),))
        conn.commit()
        conn.close()
    finally:
        wb.close()

def _prg_ensure_cache_fresh():
    mod_time = _prg_file_mod_time()
    if mod_time == 0:
        raise Exception(f"No se encontró el archivo '{PORT_RELEASE_GUIDE_PATH}'.")
    if mod_time > _prg_cached_mod_time():
        _prg_sync_cache()

# Un escaneo de site completo puede tomar varios minutos (2 workers x cientos de
# abonados). Mantener eso como una sola petición HTTP síncrona es fragil: el navegador
# o cualquier proxy/firewall intermedio puede cortar una conexión abierta e inactiva
# por varios minutos ("NetworkError when attempting to fetch resource"). Por eso se
# corre en un hilo de fondo (mismo patrón que /api/sync) y el frontend hace polling
# a /api/topology/scan/status cada pocos segundos.
topology_scan_status = {"state": "idle", "checked": 0, "total": 0, "result": None, "error": None}
topology_scan_lock = threading.Lock()

def _execute_topology_scan(site_input, olt, hubbox):
    import re
    global topology_scan_status

    site_bare = site_input.replace("OLT01", "").replace("OLT02", "")
    xb_code = f"XB0{olt}"
    hb_nums = [int(hubbox[-1])] if hubbox and hubbox != "Todas" else [1, 2, 3, 4]

    _prg_ensure_cache_fresh()
    prg_conn = sqlite3.connect(PRG_CACHE_DB)
    prg_conn.row_factory = sqlite3.Row
    prg_cursor = prg_conn.cursor()
    try:
        prg_cursor.execute("""
            SELECT n.box_code, n.account, n.department, n.customer_name, n.phone, t.bras
            FROM prg_nims n
            LEFT JOIN prg_tms t ON t.username = n.account
            WHERE n.box_code LIKE ?
        """, (f"{site_bare}-{xb_code}-%",))
        rows = prg_cursor.fetchall()
    finally:
        prg_conn.close()

    department = rows[0]["department"] if rows else ""
    branch = process_data.get_branch_from_nims(department, site_input) if rows else ""

    # Agrupar abonados activos por caja exacta (ej: "SB111", "EX111"), solo dentro del
    # XB y HB(s) pedidos. Filtrar por HB acá (no solo al armar la grilla de salida) es
    # clave: si no, un HUBBOX="HB01" seguía mandando a validar contra BRAS las cuentas
    # de las otras 3 HB del site entero, sin mostrarlas después igual.
    accounts_by_box = {}
    for r in rows:
        box = (r["box_code"] or "").upper()
        if f"-{xb_code}-" not in box:
            continue
        m = re.search(r'(SB|EB|EX)(\d{3})$', box.split("-")[-1])
        if not m:
            continue
        if int(m.group(2)[0]) not in hb_nums:
            continue
        box_key = m.group(1) + m.group(2)
        accounts_by_box.setdefault(box_key, []).append({
            "account": r["account"], "bras": r["bras"] or "",
            "customer_name": r["customer_name"] or "", "phone": r["phone"] or ""
        })

    # Validar en un solo lote (respetando el límite de 2 workers concurrentes hacia BRAS),
    # actualizando el contador de progreso a medida que cada cuenta se resuelve.
    todas_las_cuentas = [(c["account"], c["bras"]) for pairs in accounts_by_box.values() for c in pairs]
    resultados_bras = {}
    with topology_scan_lock:
        topology_scan_status["total"] = len(todas_las_cuentas)
    if todas_las_cuentas:
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            futuros = [executor.submit(_check_bras_status, acc, bras) for acc, bras in todas_las_cuentas]
            for futuro in concurrent.futures.as_completed(futuros):
                acc, estado, ip_status = futuro.result()
                resultados_bras[acc] = {"estado": estado, "ip_status": ip_status}
                with topology_scan_lock:
                    topology_scan_status["checked"] += 1

    hbs_out = []
    down_boxes = []
    total_act = total_online = total_not_online = total_unknown = total_nunca_online = 0

    for hb_num in hb_nums:
        hb_code = f"HB0{hb_num}"
        lineas_out = []
        for line_num in range(1, 5):
            boxes_teoricas = nims_topology.get_boxes_for_line(hb_num, line_num)
            cajas_out = []
            exps_out = []
            for box_code in boxes_teoricas:
                ex_code = "EX" + box_code[2:]
                for code, target_list, hide_if_empty in ((box_code, cajas_out, False), (ex_code, exps_out, True)):
                    pares = accounts_by_box.get(code, [])
                    act = len(pares)
                    clientes = [{
                        "account": c["account"], "customer_name": c["customer_name"], "phone": c["phone"],
                        "bras": c["bras"], "estado": resultados_bras.get(c["account"], {}).get("estado", "UNKNOWN")
                    } for c in pares]
                    online = sum(1 for c in clientes if c["estado"] == "ONLINE")
                    not_online = sum(1 for c in clientes if c["estado"] == "NOT ONLINE")
                    nunca_online = sum(1 for c in clientes if c["estado"] == "NUNCA ONLINE")
                    unknown = act - online - not_online - nunca_online
                    definitivos = online + not_online
                    if act == 0:
                        if hide_if_empty:
                            continue
                        color = "blue"
                        pct = 0
                    elif definitivos == 0:
                        # Ningún resultado concluyente (todo UNKNOWN/error de red tras
                        # reintento): no se puede afirmar que la caja esté caída, así que
                        # NO se colorea rojo ni se genera alarma con datos sin confirmar.
                        color = "grey"
                        pct = 0
                    else:
                        # % calculado solo sobre resultados confirmados (online+not_online),
                        # para que cuentas con estado ambiguo no distorsionen el porcentaje.
                        pct = round((online / definitivos) * 100, 1)
                        color = _color_for_pct(pct)
                        if color == "red":
                            down_boxes.append(code)
                    total_act += act
                    total_online += online
                    total_not_online += not_online
                    total_unknown += unknown
                    total_nunca_online += nunca_online
                    target_list.append({
                        "box_code": code, "act": act, "online": online,
                        "not_online": not_online, "unknown": unknown, "nunca_online": nunca_online,
                        "pct": pct, "color": color,
                        "clients": clientes
                    })
            lineas_out.append({"line": line_num, "cajas": cajas_out, "exps": exps_out})
        hbs_out.append({"hb_code": hb_code, "lineas": lineas_out})

    return {
        "success": True,
        "site": site_bare,
        "olt": olt,
        "xb_code": xb_code,
        "branch": branch,
        "hbs": hbs_out,
        "alarm": {
            "boxes_down_count": len(down_boxes),
            "boxes_down": down_boxes
        },
        "totals": {
            "act": total_act, "online": total_online, "not_online": total_not_online,
            "unknown": total_unknown, "nunca_online": total_nunca_online
        }
    }

def _run_topology_scan_bg(site_input, olt, hubbox):
    global topology_scan_status
    try:
        result = _execute_topology_scan(site_input, olt, hubbox)
        with topology_scan_lock:
            topology_scan_status["state"] = "done"
            topology_scan_status["result"] = result
    except Exception as e:
        with topology_scan_lock:
            topology_scan_status["state"] = "error"
            topology_scan_status["error"] = str(e)

@app.route("/api/topology/scan/start", methods=["POST"])
def start_topology_scan():
    global topology_scan_status
    site_input = request.args.get("site", "").strip().upper()
    olt = request.args.get("olt", "1").strip()
    hubbox = request.args.get("hubbox", "Todas").strip()

    if not site_input:
        return jsonify({"success": False, "error": "Debe especificar el parámetro 'site'."}), 400
    if olt not in ("1", "2"):
        return jsonify({"success": False, "error": "El parámetro 'olt' debe ser 1 o 2 (no existen OLT03+ en la topología real)."}), 400

    with topology_scan_lock:
        if topology_scan_status["state"] == "running":
            return jsonify({"success": False, "error": "Ya hay un escaneo de topología en curso."})
        topology_scan_status = {"state": "running", "checked": 0, "total": 0, "result": None, "error": None}

    threading.Thread(target=_run_topology_scan_bg, args=(site_input, olt, hubbox), daemon=True).start()
    return jsonify({"success": True})

@app.route("/api/topology/scan/status", methods=["GET"])
def get_topology_scan_status():
    with topology_scan_lock:
        return jsonify(dict(topology_scan_status))

# =========================================================================
# ALARMAS DE RED: escaneo periódico de TODOS los sites buscando cajas caídas.
# Escanear TODOS los clientes de TODAS las cajas (56,501 cuentas, 2 workers)
# tomaría ~9.4h, no cabe en un ciclo de 3h. Por eso el muestreo: 1 cuenta
# representativa por caja (18,519 cajas) primero; si esa muestra da NOT
# ONLINE, ENTONCES se revisan TODOS los clientes de esa caja específica para
# confirmar antes de generar la alarma (evita falsos positivos de una sola
# cuenta rara/cancelada-pero-mal-marcada-activa en NIMS).
# Aun con el muestreo, con 2 workers un ciclo completo probablemente tarde
# más de 3h en la práctica (sobre todo por los reintentos en cuentas no-
# ONLINE) — el loop no espera un reloj fijo, corre de nuevo apenas termina
# el ciclo anterior si ya pasaron 3h, así que la cadencia real será "cada
# 3h o el tiempo que tome el ciclo, lo que sea mayor", no exactamente 3h.
ALARM_SCAN_INTERVAL_SECONDS = 3 * 3600
CRITICAL_DOWN_RATIO = 0.5   # >=50% de las cajas pobladas del site caídas -> Crítica
MEDIA_MIN_BOXES_DOWN = 2    # >=2 cajas caídas (y no llega a Crítica) -> Media

def _ensure_alarms_table():
    conn = get_db_connection()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS topology_alarms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scan_time TEXT,
                site TEXT,
                branch TEXT,
                severity TEXT,
                box_code TEXT,
                act INTEGER,
                online INTEGER,
                not_online INTEGER,
                message TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alarms_scan_time ON topology_alarms(scan_time)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_alarms_severity ON topology_alarms(severity)")
        conn.commit()
    finally:
        conn.close()

alarm_scan_status = {"state": "idle", "last_run": None, "last_duration_sec": None, "checked": 0, "total": 0}
alarm_scan_lock = threading.Lock()

def _get_all_populated_boxes():
    """box_code completo -> {'site','xb','department','accounts':[(account,bras),...]}"""
    _prg_ensure_cache_fresh()
    prg_conn = sqlite3.connect(PRG_CACHE_DB)
    prg_conn.row_factory = sqlite3.Row
    cur = prg_conn.cursor()
    cur.execute("""
        SELECT n.box_code, n.account, n.department, t.bras
        FROM prg_nims n
        LEFT JOIN prg_tms t ON t.username = n.account
    """)
    rows = cur.fetchall()
    prg_conn.close()

    boxes = {}
    for r in rows:
        box = (r["box_code"] or "").upper()
        parts = box.split("-")
        if len(parts) < 4:
            continue
        entry = boxes.setdefault(box, {"site": parts[0], "xb": parts[1], "department": r["department"] or "", "accounts": []})
        entry["accounts"].append((r["account"], r["bras"] or ""))
    return boxes

def _run_alarm_scan_cycle():
    global alarm_scan_status
    start_time = time.time()
    with alarm_scan_lock:
        alarm_scan_status["state"] = "running"
        alarm_scan_status["checked"] = 0

    boxes = _get_all_populated_boxes()
    with alarm_scan_lock:
        alarm_scan_status["total"] = len(boxes)

    # Paso 1: muestreo, 1 cuenta representativa por caja (preferir una con BRAS asignado)
    samples = []
    for box_code, info in boxes.items():
        acc, bras = next((ab for ab in info["accounts"] if ab[1]), info["accounts"][0])
        samples.append((box_code, acc, bras))

    suspicious_boxes = set()
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        futuros = {executor.submit(_check_bras_status, acc, bras): box_code for box_code, acc, bras in samples}
        for futuro in concurrent.futures.as_completed(futuros):
            box_code = futuros[futuro]
            _, estado, _ = futuro.result()
            if estado != "ONLINE":
                suspicious_boxes.add(box_code)
            with alarm_scan_lock:
                alarm_scan_status["checked"] += 1

    # Paso 2: confirmar cajas sospechosas revisando TODOS sus clientes
    confirmed_down = {}
    for box_code in suspicious_boxes:
        accounts = boxes[box_code]["accounts"]
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            futuros = [executor.submit(_check_bras_status, acc, bras) for acc, bras in accounts]
            resultados = [f.result() for f in concurrent.futures.as_completed(futuros)]
        act = len(accounts)
        online = sum(1 for _, estado, _ in resultados if estado == "ONLINE")
        not_online = sum(1 for _, estado, _ in resultados if estado == "NOT ONLINE")
        definitivos = online + not_online
        if definitivos > 0 and (online / definitivos) * 100 <= 20:
            confirmed_down[box_code] = (act, online, not_online)

    # Paso 3: clasificar severidad por site (según qué fracción de SUS cajas está caída) y guardar
    down_by_site = {}
    for box_code in confirmed_down:
        down_by_site.setdefault(boxes[box_code]["site"], []).append(box_code)

    boxes_per_site = {}
    for info in boxes.values():
        boxes_per_site[info["site"]] = boxes_per_site.get(info["site"], 0) + 1

    scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db_connection()
    try:
        for site, down_boxes_list in down_by_site.items():
            total_site_boxes = boxes_per_site.get(site, len(down_boxes_list))
            down_ratio = len(down_boxes_list) / total_site_boxes if total_site_boxes else 0
            if down_ratio >= CRITICAL_DOWN_RATIO:
                severity = "critical"
            elif len(down_boxes_list) >= MEDIA_MIN_BOXES_DOWN:
                severity = "media"
            else:
                severity = "caja"

            department = boxes[down_boxes_list[0]]["department"]
            branch = process_data.get_branch_from_nims(department, site)

            for box_code in down_boxes_list:
                act, online, not_online = confirmed_down[box_code]
                conn.execute("""
                    INSERT INTO topology_alarms (scan_time, site, branch, severity, box_code, act, online, not_online, message)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (scan_time, site, branch, severity, box_code, act, online, not_online,
                      f"Caja {box_code} caída: {online}/{act} online confirmados de {len(down_boxes_list)} caja(s) caída(s) en el site"))
        conn.commit()
    finally:
        conn.close()

    with alarm_scan_lock:
        alarm_scan_status["state"] = "idle"
        alarm_scan_status["last_run"] = scan_time
        alarm_scan_status["last_duration_sec"] = round(time.time() - start_time, 1)

def _alarm_scan_loop():
    _ensure_alarms_table()
    while True:
        cycle_start = time.time()
        try:
            _run_alarm_scan_cycle()
        except Exception as e:
            print(f"[Alarm Scan Error] {e}", flush=True)
            with alarm_scan_lock:
                alarm_scan_status["state"] = "error"
        time.sleep(max(0, ALARM_SCAN_INTERVAL_SECONDS - (time.time() - cycle_start)))

CLOUD_COMMENTS_PULL_INTERVAL_SECONDS = 5 * 60

def _cloud_comments_pull_loop():
    if not cloud_sync.is_configured():
        print("[cloud_sync] CLOUD_SYNC_URL/CLOUD_API_KEY no configurados: no se van a traer comentarios de encargados.", flush=True)
        return
    while True:
        cycle_start = time.time()
        cloud_sync.pull_comments()
        time.sleep(max(0, CLOUD_COMMENTS_PULL_INTERVAL_SECONDS - (time.time() - cycle_start)))

@app.route("/api/topology/alarms", methods=["GET"])
def get_topology_alarms():
    severity = request.args.get("severity", "").strip()
    branch = request.args.get("branch", "").strip()
    site = request.args.get("site", "").strip()
    limit = min(int(request.args.get("limit", 200)), 1000)

    conn = get_db_connection()
    try:
        query = "SELECT * FROM topology_alarms WHERE 1=1"
        params = []
        if severity:
            query += " AND severity = ?"
            params.append(severity)
        if branch:
            query += " AND branch = ?"
            params.append(branch)
        if site:
            query += " AND site LIKE ?"
            params.append(f"%{site}%")
        query += " ORDER BY scan_time DESC, id DESC LIMIT ?"
        params.append(limit)
        cursor = conn.cursor()
        cursor.execute(query, params)
        rows = [dict(r) for r in cursor.fetchall()]
        return jsonify({"success": True, "alarms": rows})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()

@app.route("/api/topology/alarms/status", methods=["GET"])
def get_alarm_scan_status():
    with alarm_scan_lock:
        return jsonify(dict(alarm_scan_status))

# ==================== Credenciales (GNOC / Tableau / CNOC) ====================
# Las credenciales de estos portales cambian de tanto en tanto (ej. rotación de contraseña).
# En vez de tener que editar .env a mano cada vez, esta sección permite verlas/actualizarlas
# desde el dashboard. Nunca se devuelve la contraseña real al frontend, solo si está configurada.
ENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")

CREDENTIAL_SYSTEMS = {
    "gnoc": {"user_key": "INTRANET_USER", "pass_key": "INTRANET_PASSWORD", "label": "GNOC"},
    "tableau": {"user_key": "TABLEAU_USER", "pass_key": "TABLEAU_PASSWORD", "label": "Tableau"},
    "cnoc": {"user_key": "CNOC_USER", "pass_key": "CNOC_PASSWORD", "label": "CNOC"},
    "tms": {"user_key": "TMS_USER", "pass_key": "TMS_PASSWORD", "label": "TMs"},
    "nims": {"user_key": "NIMS_USER", "pass_key": "NIMS_PASSWORD", "label": "NIMS"},
}

def read_env_dict():
    env = {}
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, "r", encoding="utf-8") as f:
            for line in f:
                stripped = line.strip()
                if not stripped or stripped.startswith("#") or "=" not in stripped:
                    continue
                k, v = stripped.split("=", 1)
                env[k.strip()] = v.strip()
    return env

def update_env_file(updates):
    """Actualiza (o agrega) pares KEY=VALUE en .env preservando el resto del archivo tal cual
    (comentarios, orden, líneas en blanco). No toca ninguna línea que no esté en `updates`."""
    lines = []
    if os.path.exists(ENV_PATH):
        with open(ENV_PATH, "r", encoding="utf-8") as f:
            lines = f.readlines()

    remaining = dict(updates)
    new_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped and not stripped.startswith("#") and "=" in stripped:
            key = stripped.split("=", 1)[0].strip()
            if key in remaining:
                new_lines.append(f"{key}={remaining.pop(key)}\n")
                continue
        new_lines.append(line)

    for k, v in remaining.items():
        new_lines.append(f"{k}={v}\n")

    with open(ENV_PATH, "w", encoding="utf-8") as f:
        f.writelines(new_lines)

@app.route("/api/settings/credentials", methods=["GET"])
def get_credentials():
    env = read_env_dict()
    result = {}
    for key, cfg in CREDENTIAL_SYSTEMS.items():
        result[key] = {
            "label": cfg["label"],
            "username": env.get(cfg["user_key"], ""),
            "has_password": bool(env.get(cfg["pass_key"], "")),
        }
    return jsonify(result)

@app.route("/api/settings/credentials", methods=["POST"])
def update_credentials():
    with sync_lock:
        if sync_status["state"] in ("downloading", "processing"):
            return jsonify({"success": False, "message": "No se pueden cambiar credenciales mientras hay una sincronización activa."})

    body = request.get_json(silent=True) or {}
    system = body.get("system")
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""

    if system not in CREDENTIAL_SYSTEMS:
        return jsonify({"success": False, "message": "Sistema desconocido."}), 400
    if not username and not password:
        return jsonify({"success": False, "message": "Debes indicar un usuario y/o una contraseña nuevos."}), 400

    cfg = CREDENTIAL_SYSTEMS[system]
    updates = {}
    if username:
        updates[cfg["user_key"]] = username
    if password:
        updates[cfg["pass_key"]] = password

    try:
        update_env_file(updates)
    except Exception as e:
        return jsonify({"success": False, "message": f"Error al guardar en .env: {e}"}), 500

    return jsonify({"success": True, "message": f"Credenciales de {cfg['label']} actualizadas. Se usarán en la próxima sincronización."})

# Endpoint: Sincronizar Excel con Base de Datos
@app.route("/api/sync", methods=["POST"])
def sync_data():
    global sync_status
    with sync_lock:
        if sync_status["state"] in ("downloading", "processing"):
            return jsonify({"success": False, "message": "Ya hay una sincronización activa en curso."})

    gnoc_env_overrides = None
    body = request.get_json(silent=True) or {}
    from_month = (body.get("from_month") or "").strip()
    to_month = (body.get("to_month") or "").strip()
    if from_month or to_month:
        if not (from_month and to_month):
            return jsonify({"success": False, "message": "Debes indicar tanto el mes de inicio como el de fin."})
        try:
            filter_create_time = compute_filter_create_time_for_months(from_month, to_month)
        except ValueError:
            return jsonify({"success": False, "message": "Formato de mes inválido (se espera YYYY-MM)."})
        if from_month > to_month:
            return jsonify({"success": False, "message": "El mes de inicio no puede ser posterior al mes de fin."})
        gnoc_env_overrides = {"FILTER_CREATE_TIME": filter_create_time}
        save_sync_month_range(from_month, to_month)
    else:
        # "Automático" explícito (ambos selects vacíos): el ciclo automático debe volver a
        # usar el FILTER_CREATE_TIME de .env en vez de seguir repitiendo el último rango elegido.
        clear_sync_month_range()

    threading.Thread(target=run_background_sync, args=(gnoc_env_overrides,), daemon=True).start()
    return jsonify({"success": True, "message": "Sincronización iniciada en segundo plano."})

# Endpoint: Consultar Estado de la Sincronización
@app.route("/api/sync/status", methods=["GET"])
def get_sync_status():
    global sync_status
    with sync_lock:
        status_copy = dict(sync_status)
    return jsonify(status_copy)

# Endpoint: Cancelar / Desbloquear Sincronización atascada
@app.route("/api/sync/cancel", methods=["POST"])
def cancel_sync():
    global sync_status
    with sync_lock:
        sync_status["state"] = "idle"
        sync_status["message"] = ""
        for proc in list(active_sync_processes):
            try:
                proc.terminate()
            except Exception:
                pass
        active_sync_processes.clear()
    return jsonify({"success": True, "message": "Sincronización cancelada. Puedes iniciar una nueva."})

# Endpoint: Obtener Estadísticas y Métricas consolidadas
@app.route("/api/stats", methods=["GET"])
def get_stats():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Total orders (incluyendo cerradas y activas)
        cursor.execute("SELECT COUNT(*) FROM work_orders")
        total_orders = cursor.fetchone()[0]
        
        # Total errors (vtp_marlo.delacruz)
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 1")
        total_errors = cursor.fetchone()[0]
        
        # Valid active pending WOs (excluyendo cerradas)
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed')")
        total_valid = cursor.fetchone()[0]
        
        # Pending intervals for valid active orders
        # < 24h
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed') AND pending_hours < 24")
        pending_24 = cursor.fetchone()[0]
        
        # 24h - 48h
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed') AND pending_hours >= 24 AND pending_hours < 48")
        pending_48 = cursor.fetchone()[0]
        
        # 48h - 72h
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed') AND pending_hours >= 48 AND pending_hours < 72")
        pending_72 = cursor.fetchone()[0]
        
        # > 72h
        cursor.execute("SELECT COUNT(*) FROM work_orders WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed') AND pending_hours >= 72")
        pending_older = cursor.fetchone()[0]
        
        # Distribución por Close Reason (excluyendo Pendiente/Error)
        cursor.execute("""
            SELECT close_reason, COUNT(*) as count 
            FROM work_orders 
            WHERE close_reason NOT IN ('Pendiente (Sin comentario)', 'Error (vtp_marlo.delacruz)')
            GROUP BY close_reason 
            ORDER BY count DESC
        """)
        reasons = [{"reason": row["close_reason"], "count": row["count"]} for row in cursor.fetchall()]
        
        # Distribución por CD Group (Equipo de soporte)
        cursor.execute("""
            SELECT cd_group, COUNT(*) as count 
            FROM work_orders 
            GROUP BY cd_group 
            ORDER BY count DESC
        """)
        cd_groups = [{"cd_group": row["cd_group"], "count": row["count"]} for row in cursor.fetchall()]

        # Distribución por Branch (averías PENDIENTES, mismo filtro que total_valid/pending_intervals
        # arriba -- excluye errores de Marlo y WOs ya cerradas, no es el total histórico)
        cursor.execute("""
            SELECT COALESCE(NULLIF(branch, ''), 'SIN BRANCH') as branch_name, COUNT(*) as count
            FROM work_orders
            WHERE is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed')
            GROUP BY branch_name
            ORDER BY count DESC
        """)
        branches = [{"branch": row["branch_name"], "count": row["count"]} for row in cursor.fetchall()]

        # Fecha/hora de la última sincronización de Excel exitosa (ver process_data.py).
        # sync_meta puede no existir todavía en bases de datos que no pasaron por el
        # db_setup.py más reciente, así que se tolera su ausencia.
        last_sync_at = None
        try:
            cursor.execute("SELECT value FROM sync_meta WHERE key = 'last_sync_at'")
            row = cursor.fetchone()
            last_sync_at = row["value"] if row else None
        except sqlite3.OperationalError:
            pass

        return jsonify({
            "total_orders": total_orders,
            "total_errors": total_errors,
            "total_valid": total_valid,
            "last_sync_at": last_sync_at,
            "pending_intervals": {
                "under_24h": pending_24,
                "under_48h": pending_48,
                "under_72h": pending_72,
                "over_72h": pending_older
            },
            "reasons_distribution": reasons,
            "cd_groups_distribution": cd_groups,
            "branch_distribution": branches
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

# Endpoint: Listado completo de órdenes de trabajo (para la grilla/tabla)
@app.route("/api/work_orders", methods=["GET"])
def get_work_orders():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT wo_code, wo_name, wo_type, description, wo_status, 
                   create_time, cd_group, ft_technician, priority, 
                   ft_comment, pending_hours, close_reason, is_error,
                   branch, ticket_code, wo_create_date, responsible_unit, account,
                   warranty_period, implementation_test, act_status, sub_status,
                   online_status, ft_gnoc, staf_team, connector_code, compcontent,
                   closed_time, resolution_hours
            FROM work_orders
            ORDER BY pending_hours DESC
        """)
        orders = []
        for row in cursor.fetchall():
            orders.append({
                "wo_code": row["wo_code"],
                "wo_name": row["wo_name"],
                "wo_type": row["wo_type"],
                "description": row["description"],
                "wo_status": row["wo_status"],
                "create_time": row["create_time"],
                "cd_group": row["cd_group"],
                "ft_technician": row["ft_technician"],
                "priority": row["priority"],
                "ft_comment": row["ft_comment"],
                "pending_hours": round(row["pending_hours"], 2),
                "close_reason": row["close_reason"],
                "is_error": bool(row["is_error"]),
                "branch": row["branch"] or "",
                "ticket_code": row["ticket_code"] or "",
                "wo_create_date": row["wo_create_date"] or "",
                "responsible_unit": row["responsible_unit"] or "",
                "account": row["account"] or "",
                "warranty_period": row["warranty_period"] or "",
                "implementation_test": row["implementation_test"] or "",
                "act_status": row["act_status"] or "",
                "sub_status": row["sub_status"] or "",
                "online_status": row["online_status"] or "",
                "ft_gnoc": row["ft_gnoc"] or "",
                "staf_team": row["staf_team"] or "",
                "connector_code": row["connector_code"] or "",
                "compcontent": row["compcontent"] or "",
                "closed_time": row["closed_time"] or "",
                "resolution_hours": round(row["resolution_hours"], 2) if row["resolution_hours"] is not None else None
            })
        return jsonify(orders)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


# Endpoint: Exportar órdenes de trabajo a Excel con formato exacto de 16 columnas
@app.route("/api/work_orders/export", methods=["GET"])
@app.route("/api/work_orders/export/<filename>", methods=["GET"])
def export_work_orders(filename=None):
    search_query = request.args.get("search", "").strip().lower()
    filter_type = request.args.get("type", "all")
    filter_cd = request.args.get("cd", "all")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT branch, wo_code, ticket_code, wo_create_date, responsible_unit, 
                   ft_technician, account, warranty_period, implementation_test, 
                   act_status, sub_status, online_status, ft_gnoc, staf_team, 
                   connector_code, compcontent, wo_name, description, wo_status, 
                   close_reason, is_error, cd_group, pending_hours, create_time
            FROM work_orders
        """
        where_clauses = []
        params = []
        
        # Filtro de tipo
        if filter_type == "pending":
            where_clauses.append("LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed')")
        elif filter_type == "valid":
            where_clauses.append("is_error = 0 AND LOWER(wo_status) NOT IN ('close', 'closed', 'closed ft', 'ft completed')")
        elif filter_type == "error":
            where_clauses.append("is_error = 1")
            
        # Filtro de CD
        if filter_cd != "all" and filter_cd:
            where_clauses.append("cd_group = ?")
            params.append(filter_cd)
            
        # Filtro de búsqueda textual (mismo criterio de filtrado frontend)
        if search_query:
            where_clauses.append("""
                (LOWER(wo_code) LIKE ? OR 
                 LOWER(ft_technician) LIKE ? OR 
                 LOWER(cd_group) LIKE ? OR 
                 LOWER(close_reason) LIKE ? OR 
                 LOWER(wo_name) LIKE ? OR
                 LOWER(branch) LIKE ? OR
                 LOWER(connector_code) LIKE ?)
            """)
            like_param = f"%{search_query}%"
            params.extend([like_param] * 7)
            
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
            
        query += " ORDER BY pending_hours DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Crear libro de Excel en memoria
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Work Orders"
        
        # Encabezados en el orden exacto de Tableau (remplazando FT Reassigned por FT de GNOC y agregando Estado WO al final)
        headers = [
            "BRANCH", "WO_CODE", "TICKET_CODE", "WO_CREATE_DATE", 
            "RESPONSIBLE_UNIT_WO", "FT Reassigned", "Account", 
            "Warranty Days", "Implementation test", "Act Status", 
            "Sub Status", "Online Status", "FT Gnoc", "Staf Team", 
            "Connector_Code", "COMPCONTENT", "Estado WO"
        ]
        ws.append(headers)
        
        # Estilo de encabezados (Premium Blue)
        from openpyxl.styles import Font, PatternFill
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            
        # Datos
        for row in rows:
            ws.append([
                row["branch"] or "",
                row["wo_code"] or "",
                row["ticket_code"] or "",
                row["wo_create_date"] or row["create_time"] or "",
                row["responsible_unit"] or row["cd_group"] or "",
                row["ft_technician"] or "", # FT Reassigned se llena obligatoriamente con el FT de GNOC
                row["account"] or "",
                row["warranty_period"] or "",
                row["implementation_test"] or "",
                row["act_status"] or "",
                row["sub_status"] or "",
                row["online_status"] or "",
                row["ft_gnoc"] or "",
                row["staf_team"] or "",
                row["connector_code"] or "",
                row["compcontent"] or "",
                row["wo_status"] or ""
            ])
            
        # Ajuste automático del ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
            
        # Guardar en stream BytesIO
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = "reporte_ordenes_trabajo.xlsx" if not search_query and filter_type == "all" and filter_cd == "all" else "reporte_ordenes_filtradas.xlsx"
        
        return send_file(
            file_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

# ─────────────────────────────────────────────────────────────────────────────
#  OLT AUDITOR — Rutas de la API
# ─────────────────────────────────────────────────────────────────────────────
_olt_scan_thread = None

@app.route("/api/olt/status")
def olt_status():
    return jsonify(olt_audit.get_status())


@app.route("/api/olt/list", methods=["GET"])
def olt_list():
    import pandas as pd
    if not os.path.exists(OLT_INPUT):
        return jsonify([])
    try:
        df = pd.read_excel(OLT_INPUT)
        if 'OLT_NAME' not in df.columns or 'OLT_IP' not in df.columns:
            return jsonify([])
        olts = [{"name": row['OLT_NAME'], "ip": row['OLT_IP']} for _, row in df.iterrows()]
        return jsonify(olts)
    except Exception as e:
        print(f"[OLT List] Error: {e}")
        return jsonify([])

@app.route("/api/olt/ems_status")
def olt_ems_status():
    """Verifica si el servidor EMS es alcanzable vía TCP antes de iniciar un escaneo."""
    import socket as _socket
    from urllib.parse import urlparse
    ems_url = olt_audit.EMS_URL
    parsed = urlparse(ems_url)
    host = parsed.hostname
    port = parsed.port or 80
    try:
        sock = _socket.create_connection((host, port), timeout=8)
        sock.close()
        return jsonify({"ok": True, "host": host, "port": port,
                        "message": f"EMS alcanzable en {host}:{port}"})
    except (_socket.timeout, OSError) as e:
        return jsonify({"ok": False, "host": host, "port": port,
                        "message": f"EMS NO alcanzable ({host}:{port}): {str(e)[:80]}"}), 503


@app.route("/api/olt/scan", methods=["POST"])
def olt_scan():
    global _olt_scan_thread
    status = olt_audit.get_status()
    if status["state"] == "scanning":
        return jsonify({"error": "Ya hay un escaneo en curso."}), 409
    if not os.path.exists(OLT_INPUT):
        return jsonify({
            "error": f"No se encontró {OLT_INPUT}. "
                     "Crea el archivo con columnas OLT_NAME y OLT_IP."
        }), 400

    data = request.get_json(silent=True) or {}
    selected_olts = data.get("selected_olts", [])

    def _run():
        try:
            olt_audit.run_audit(OLT_INPUT, selected_olts)
        except Exception as e:
            print(f"[OLT Scan] Error: {e}", flush=True)

    _olt_scan_thread = threading.Thread(target=_run, daemon=True)
    _olt_scan_thread.start()
    return jsonify({"ok": True, "message": "Escaneo OLT iniciado."})


@app.route("/api/olt/scan/cancel", methods=["POST"])
def olt_scan_cancel():
    olt_audit.cancel_audit()
    return jsonify({"ok": True, "message": "Cancelación solicitada."})


@app.route("/api/olt/fallas")
def olt_fallas():
    filtro_olt  = request.args.get("olt", "")
    filtro_tipo = request.args.get("tipo", "")
    filtro_prio = request.args.get("prioridad", "")
    limit       = int(request.args.get("limit", 500))
    rows = olt_audit.get_fallas(filtro_olt, filtro_tipo, filtro_prio, limit)
    return jsonify(rows)


@app.route("/api/olt/cortes")
def olt_cortes():
    import re
    rows = olt_audit.get_cortes()

    # Branch por OLT: se deriva el site_code quitando el sufijo "OLTxx" del nombre (ej.
    # "ARE0005OLT01" -> "ARE0005") y se busca en kpi_zonas_map (mismo mapeo ZONAS que ya
    # usa el Reporte KPI para resolver branch por site_code, ver kpi_calc.py).
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT site_code, branch FROM kpi_zonas_map")
    zonas_map = {r["site_code"].upper(): r["branch"] for r in cursor.fetchall()}
    conn.close()

    for row in rows:
        site_code = re.sub(r'OLT\d+$', '', row["olt_name"] or "", flags=re.IGNORECASE).upper()
        row["branch"] = zonas_map.get(site_code, "")

    return jsonify(rows)


@app.route("/api/olt/alarmas")
def olt_alarmas():
    rows = olt_audit.get_alarmas()
    return jsonify(rows)


@app.route("/api/olt/resumen")
def olt_resumen():
    rows = olt_audit.get_resumen_olts()
    return jsonify(rows)


@app.route("/api/olt/errors")
def olt_errors():
    rows = olt_audit.get_olt_errors()
    return jsonify(rows)


# Búsqueda de cliente: reemplaza al viejo listado de "Puertos Caídos" (que mostraba TODAS
# las ONUs en falla activa, poco accionable). Busca por cuenta (exacto), o por teléfono/
# nombre/cuenta parcial en NIMS, resuelve el puerto TMS del cliente y devuelve el estado
# de su ONU en la OLT -mismo parseo de 'OLTNAME/CONST/PUERTO/ID[:extra]' que ya usa el
# botón "Estado del Puerto" en el detalle de WO (runOltPortStatus en app.js).
@app.route("/api/olt/client_search")
def olt_client_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "Ingresa una cuenta, teléfono o nombre de cliente."}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT username FROM tm_subscribers WHERE username = ? COLLATE NOCASE", (q,))
        row = cursor.fetchone()
        account = row["username"] if row else None

        if not account:
            like = f"%{q}%"
            cursor.execute("""
                SELECT DISTINCT account, customer_name, phone
                FROM nims_subscribers
                WHERE account LIKE ? OR phone LIKE ? OR customer_name LIKE ?
                LIMIT 15
            """, (like, like, like))
            matches = cursor.fetchall()
            if len(matches) == 0:
                return jsonify({"error": f"No se encontró ningún cliente que coincida con \"{q}\"."}), 404
            if len(matches) > 1:
                return jsonify({
                    "multiple": True,
                    "candidates": [
                        {"account": r["account"], "customer_name": r["customer_name"], "phone": r["phone"]}
                        for r in matches
                    ]
                })
            account = matches[0]["account"]

        cursor.execute("SELECT username, port, mac, status, bras FROM tm_subscribers WHERE username = ? COLLATE NOCASE", (account,))
        tm_row = cursor.fetchone()
        cursor.execute("""
            SELECT customer_name, phone, address, box_code, connector_code, site_name, status
            FROM nims_subscribers WHERE account = ? COLLATE NOCASE LIMIT 1
        """, (account,))
        nims_row = cursor.fetchone()

        base = {
            "account": account,
            "customer_name": nims_row["customer_name"] if nims_row else None,
            "phone": nims_row["phone"] if nims_row else None,
            "address": nims_row["address"] if nims_row else None,
            "box_code": nims_row["box_code"] if nims_row else None,
            "tms_status": tm_row["status"] if tm_row else None,
        }

        if not tm_row or not tm_row["port"]:
            base["error"] = "El cliente no tiene un puerto TMS registrado; no se puede consultar la OLT."
            return jsonify(base)

        parts = tm_row["port"].split("/")
        if len(parts) < 4:
            base["error"] = f"Formato de puerto TMS no reconocido: \"{tm_row['port']}\""
            return jsonify(base)

        olt_name = parts[0]
        puerto = parts[-2]
        onu_id = parts[-1].split(":")[0]
        base["tms_port_raw"] = tm_row["port"]

        # Consulta EN VIVO al EMS (1 puerto, ~1-3s) en vez de esperar a que el loop
        # continuo (400 OLTs) llegue a esta OLT -el cliente quiere ver su estado ACTUAL,
        # no el de hace potencialmente varios minutos. Si por algo falla (ej. la OLT nunca
        # se escaneó y no se conoce su IP), se cae al último dato cacheado como respaldo.
        olt_status = olt_audit.get_olt_port_status_live(olt_name, puerto, onu_id)
        if not olt_status:
            olt_status = olt_audit.get_olt_port_status(olt_name, puerto, onu_id)
        base["olt_status"] = olt_status
        if not olt_status:
            base["error"] = f"No hay datos de escaneo para {olt_name} puerto {puerto} todavía."
        return jsonify(base)
    finally:
        conn.close()


@app.route("/api/olt/detail")
def olt_detail():
    olt_name = request.args.get("olt_name", "").strip()
    if not olt_name:
        return jsonify({"error": "Parámetro olt_name requerido."}), 400
    detail = olt_audit.get_olt_detail(olt_name)
    if not detail:
        return jsonify({"error": f"No se encontraron datos para la OLT {olt_name}."}), 404
    return jsonify(detail)


# Estado de un puerto PON específico (1-16) de una OLT. Consulta el EMS EN VIVO para ese
# único puerto (rápido: 1 llamada SOAP, no las 6400 del ciclo completo), así que responde
# con el estado actual real del cliente en vez del último dato del loop continuo.
@app.route("/api/olt/port_status")
def olt_port_status():
    olt_name = request.args.get("olt_name", "").strip()
    port = request.args.get("port", "").strip()
    onu_id = request.args.get("onu_id", "").strip()
    if not olt_name or not port:
        return jsonify({"error": "Parámetros 'olt_name' y 'port' requeridos."}), 400
    result = olt_audit.get_olt_port_status_live(olt_name, port, onu_id or None)
    if not result:
        result = olt_audit.get_olt_port_status(olt_name, port, onu_id or None)
    if not result:
        return jsonify({"error": f"Sin datos de escaneo para {olt_name} puerto {port} todavía. El ciclo de escaneo continuo puede no haber llegado a ese puerto aún."}), 404
    return jsonify({"success": True, **result})


# ─────────────────────────────────────────────────────────────────────────────
#  KPI — WO Incident Report (replica de la hoja "KPI 2026" calculada en Python
#  desde work_orders, ver kpi_calc.py)
# ─────────────────────────────────────────────────────────────────────────────
def _sanitize_for_json(obj):
    """Convierte tipos numpy/pandas (float64, int64, NaN, NaT, Timestamp) a tipos nativos
    de Python antes de jsonify, que no sabe serializar los primeros dos."""
    import math
    import numpy as np
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_for_json(v) for v in obj]
    if isinstance(obj, (np.floating,)):
        f = float(obj)
        return None if math.isnan(f) else f
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, float) and math.isnan(obj):
        return None
    return obj


@app.route("/api/kpi/periods")
def kpi_periods():
    try:
        return jsonify(_sanitize_for_json(kpi_calc.get_available_periods()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/monthly")
def kpi_monthly():
    month = request.args.get("month", "").strip()
    if not month:
        return jsonify({"error": "Parámetro month requerido (YYYY-MM)."}), 400
    try:
        return jsonify(_sanitize_for_json(kpi_calc.compute_monthly_kpis(month)))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/weekly")
def kpi_weekly():
    week = request.args.get("week", "").strip()
    if not week:
        return jsonify({"error": "Parámetro week requerido (YYYY-MM-DD, lunes de la semana)."}), 400
    try:
        return jsonify(_sanitize_for_json(kpi_calc.compute_weekly_kpis(week)))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/trend")
def kpi_trend():
    try:
        return jsonify(_sanitize_for_json(kpi_calc.compute_trend()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/weekly_trend")
def kpi_weekly_trend():
    try:
        return jsonify(_sanitize_for_json(kpi_calc.compute_weekly_trend()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/reference_status")
def kpi_reference_status():
    try:
        return jsonify(_sanitize_for_json(kpi_calc.get_reference_status()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/kpi/refresh_reference", methods=["POST"])
def kpi_refresh_reference():
    try:
        return jsonify(_sanitize_for_json(kpi_calc.refresh_reference_data()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  REPORTE DIARIO — instalaciones (deployments) por día/semana por branch, y
#  averías pendientes/cerradas según GNOC (work_orders), desde julio en adelante.
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/daily_report/overview")
def daily_report_overview():
    try:
        return jsonify(_sanitize_for_json({
            "pending": daily_report.compute_pending_by_branch(),
            "closures_by_month": daily_report.compute_closures_by_month(),
            "closures_by_week": daily_report.compute_closures_by_week(),
            "closures_by_day": daily_report.compute_closures_by_day(),
            "installs_periods": daily_report.get_installs_available_periods(),
        }))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/daily_report/installs")
def daily_report_installs():
    try:
        period_type = request.args.get("type", "month")
        period_key = request.args.get("key", "").strip()
        if not period_key:
            return jsonify({"error": "Falta el parámetro 'key' (mes o semana)."}), 400
        if period_type == "week":
            data = daily_report.compute_installs_weekly(period_key)
        else:
            data = daily_report.compute_installs_daily(period_key)
        return jsonify(_sanitize_for_json({"type": period_type, "key": period_key, "data": data}))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  DESPLIEGUES PENDIENTES — integra la automatización "deploy ant" (Tableau
#  Deploy WO Pending -> Google Sheet propio) al dashboard.
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/deploy_pending/summary")
def deploy_pending_summary():
    try:
        data = deploy_pending.get_summary()
        comments = cloud_sync.get_cached_comments()
        for c in data.get("clients", []):
            info = comments.get(c.get("account"))
            c["comment"] = info.get("comment", "") if info else ""
            c["status"] = info.get("status", "") if info else ""
            c["comment_updated_by"] = info.get("comment_updated_by", "") if info else ""
            c["comment_updated_at"] = info.get("comment_updated_at") if info else None
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/deploy_pending/run", methods=["POST"])
def deploy_pending_run():
    started = deploy_pending.trigger_run()
    if not started:
        return jsonify({"error": "Ya hay una actualización de despliegues en curso."}), 409
    return jsonify({"ok": True, "message": "Actualización de despliegues iniciada."})


@app.route("/api/deploy_pending/run_status")
def deploy_pending_run_status():
    return jsonify(deploy_pending.get_run_status())


@app.route("/api/deploy_pending/cancel", methods=["POST"])
def deploy_pending_cancel():
    cancelled = deploy_pending.cancel_run()
    if not cancelled:
        return jsonify({"error": "No hay ninguna actualización de despliegues en curso."}), 409
    return jsonify({"ok": True, "message": "Actualización cancelada."})


# ─────────────────────────────────────────────────────────────────────────────
#  OLT AUDITOR — Loop automático de escaneo continuo 24/7
# ─────────────────────────────────────────────────────────────────────────────
OLT_SCAN_PAUSE_SEC = 5       # Pausa mínima entre ciclos (5 seg) para sincronizar I/O

_olt_loop_lock  = threading.Lock()
_olt_loop_state = {
    "running":       False,
    "last_scan":     None,    # Timestamp del último escaneo completado
    "next_scan":     "Inmediato (Modo Continuo)",
    "cycle":         0,       # Ciclos completados
    "input_missing": False,   # True si falta olts_input.xlsx
}


def _get_olt_loop_state():
    with _olt_loop_lock:
        return dict(_olt_loop_state)


def _olt_scan_loop():
    """Loop daemon: ejecuta run_audit de forma ininterrumpida y continua (ciclo tras ciclo)."""
    import time as _time
    import datetime as _dt

    print("[OLT Loop] Loop continuo iniciado — escaneando ininterrumpidamente 24/7 ciclo tras ciclo.", flush=True)

    while True:
        # Verificar que existe el archivo de entrada
        if not os.path.exists(OLT_INPUT):
            with _olt_loop_lock:
                _olt_loop_state["input_missing"] = True
                _olt_loop_state["running"] = False
            print(f"[OLT Loop] {OLT_INPUT} no encontrado — reintentando en 15 s.", flush=True)
            _time.sleep(15)
            continue

        with _olt_loop_lock:
            _olt_loop_state["input_missing"] = False

        # Esperar si ya hay un escaneo manual en curso
        current = olt_audit.get_status()
        if current["state"] == "scanning":
            print("[OLT Loop] Escaneo en curso — esperando 5 s.", flush=True)
            _time.sleep(5)
            continue

        with _olt_loop_lock:
            _olt_loop_state["running"] = True
            cycle_num = _olt_loop_state["cycle"] + 1

        ts_inicio = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[OLT Loop] Ciclo #{cycle_num} iniciado a las {ts_inicio}", flush=True)

        try:
            olt_audit.run_audit(OLT_INPUT)
            with _olt_loop_lock:
                _olt_loop_state["last_scan"] = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                _olt_loop_state["cycle"] = cycle_num
        except Exception as e:
            print(f"[OLT Loop] Error en ciclo #{cycle_num}: {e}", flush=True)

        with _olt_loop_lock:
            _olt_loop_state["running"] = False
            _olt_loop_state["next_scan"] = "Inmediato (Modo Continuo)"

        print(f"[OLT Loop] Ciclo #{cycle_num} completado. Iniciando siguiente ciclo inmediatamente...", flush=True)

        # Pausa mínima de 5 s antes de reiniciar inmediatamente el siguiente ciclo
        _time.sleep(OLT_SCAN_PAUSE_SEC)


@app.route("/api/olt/loop_status")
def olt_loop_status():
    """Estado del loop automático + estado del escaneo + info del archivo de entrada."""
    loop = _get_olt_loop_state()
    scan = olt_audit.get_status()
    return jsonify({
        "loop":         loop,
        "scan":         scan,
        "input_exists": os.path.exists(OLT_INPUT),
        "continuous":   True,
        "pause_sec":    OLT_SCAN_PAUSE_SEC,
    })


# ============================================================
# FBB DATA — rutas del módulo integrado desde DataBaseFBB
# ============================================================

@app.route("/api/fbb/dashboard", methods=["GET"])
def fbb_get_dashboard():
    try:
        branch = request.args.get("branch", "").strip()
        zone = request.args.get("zone", "").strip()
        stats = FBBManager.get_dashboard_stats(branch=branch, zone=zone)
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/filters", methods=["GET"])
def fbb_get_filters():
    try:
        filters = FBBManager.get_filter_options()
        return jsonify(filters)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/branches/<branch>/zones", methods=["GET"])
def fbb_get_branch_zones(branch):
    try:
        conn = FBBManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT zone FROM zones WHERE branch = ? AND zone IS NOT NULL AND zone != '' ORDER BY zone", (branch,))
        zones = [r["zone"] for r in cursor.fetchall()]
        conn.close()
        return jsonify(zones)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones", methods=["GET"])
def fbb_get_zones():
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 15))
        search = request.args.get("search", "").strip()
        branch = request.args.get("branch", "").strip()
        department = request.args.get("department", "").strip()
        sort_by = request.args.get("sort_by", "zone").strip()
        sort_dir = request.args.get("sort_dir", "ASC").strip()
        result = FBBManager.get_zones(
            page=page, per_page=per_page, search=search, branch=branch,
            department=department, sort_by=sort_by, sort_dir=sort_dir
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones/<int:zone_id>", methods=["GET"])
def fbb_get_zone_detail(zone_id):
    try:
        zone = FBBManager.get_zone(zone_id)
        if zone:
            return jsonify(zone)
        return jsonify({"error": "Zona no encontrada"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones", methods=["POST"])
def fbb_create_zone():
    try:
        data = request.json
        if not data or not data.get("zone"):
            return jsonify({"error": "El campo 'zone' (nombre de la zona) es obligatorio"}), 400
        new_id = FBBManager.add_zone(data)
        return jsonify({"message": "Zona creada exitosamente", "id": new_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones/<int:zone_id>", methods=["PUT"])
def fbb_update_zone(zone_id):
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No se proporcionaron datos para actualizar"}), 400
        success = FBBManager.update_zone(zone_id, data)
        if success:
            return jsonify({"message": "Zona actualizada exitosamente"})
        return jsonify({"error": "Zona no encontrada o no se realizaron cambios"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones/<int:zone_id>", methods=["DELETE"])
def fbb_delete_zone(zone_id):
    try:
        success = FBBManager.delete_zone(zone_id)
        if success:
            return jsonify({"message": "Zona eliminada exitosamente"})
        return jsonify({"error": "Zona no encontrada"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes", methods=["GET"])
def fbb_get_boxes():
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 15))
        search = request.args.get("search", "").strip()
        zone = request.args.get("zone", "").strip()
        olt = request.args.get("olt", "").strip()
        branch = request.args.get("branch", "").strip()
        box_class = request.args.get("box_class", "").strip()
        site_logical = request.args.get("site_logical", "").strip()
        sort_by = request.args.get("sort_by", "node_code").strip()
        sort_dir = request.args.get("sort_dir", "ASC").strip()
        result = FBBManager.get_boxes(
            page=page, per_page=per_page, search=search, zone=zone, olt=olt,
            branch=branch, box_class=box_class, site_logical=site_logical,
            sort_by=sort_by, sort_dir=sort_dir
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes/<int:box_id>", methods=["GET"])
def fbb_get_box_detail(box_id):
    try:
        box = FBBManager.get_box(box_id)
        if box:
            return jsonify(box)
        return jsonify({"error": "Caja no encontrada"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes", methods=["POST"])
def fbb_create_box():
    try:
        data = request.json
        if not data or not data.get("node_code"):
            return jsonify({"error": "El campo 'node_code' (código de nodo) es obligatorio"}), 400
        new_id = FBBManager.add_box(data)
        return jsonify({"message": "Caja creada exitosamente", "id": new_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes/<int:box_id>", methods=["PUT"])
def fbb_update_box(box_id):
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No se proporcionaron datos para actualizar"}), 400
        success = FBBManager.update_box(box_id, data)
        if success:
            return jsonify({"message": "Caja actualizada exitosamente"})
        return jsonify({"error": "Caja no encontrada o no se realizaron cambios"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes/<int:box_id>", methods=["DELETE"])
def fbb_delete_box(box_id):
    try:
        success = FBBManager.delete_box(box_id)
        if success:
            return jsonify({"message": "Caja eliminada exitosamente"})
        return jsonify({"error": "Caja no encontrada"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/boxes/nearest", methods=["GET"])
def fbb_get_nearest_boxes():
    try:
        lat = float(request.args.get("latitude"))
        lng = float(request.args.get("longitude"))
        limit = int(request.args.get("limit", 12))
        result = FBBManager.get_nearest_boxes(lat, lng, limit)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/fbb/staff", methods=["GET"])
def fbb_get_staff():
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 15))
        search = request.args.get("search", "").strip()
        branch = request.args.get("branch", "").strip()
        partner = request.args.get("partner", "").strip()
        sort_by = request.args.get("sort_by", "staff_team").strip()
        sort_dir = request.args.get("sort_dir", "ASC").strip()
        result = FBBManager.get_staff(
            page=page, per_page=per_page, search=search, branch=branch,
            partner=partner, sort_by=sort_by, sort_dir=sort_dir
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff/<int:staff_id>", methods=["GET"])
def fbb_get_staff_detail(staff_id):
    try:
        member = FBBManager.get_staff_member(staff_id)
        if member:
            return jsonify(member)
        return jsonify({"error": "Miembro del personal no encontrado"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff", methods=["POST"])
def fbb_create_staff():
    try:
        data = request.json
        if not data or not data.get("staff_team") or not data.get("zone"):
            return jsonify({"error": "Los campos 'staff_team' y 'zone' son obligatorios"}), 400
        new_id = FBBManager.add_staff(data)
        return jsonify({"message": "Miembro del personal creado exitosamente", "id": new_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff/<int:staff_id>", methods=["PUT"])
def fbb_update_staff(staff_id):
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No se proporcionaron datos para actualizar"}), 400
        success = FBBManager.update_staff(staff_id, data)
        if success:
            return jsonify({"message": "Información del personal actualizada exitosamente"})
        return jsonify({"error": "Personal no encontrado o no se realizaron cambios"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff/<int:staff_id>", methods=["DELETE"])
def fbb_delete_staff(staff_id):
    try:
        success = FBBManager.delete_staff(staff_id)
        if success:
            return jsonify({"message": "Registro de personal eliminado exitosamente"})
        return jsonify({"error": "Personal no encontrado"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/sync", methods=["POST"])
def fbb_sync_with_google_sheet():
    try:
        zones_count, boxes_count, staff_count, incidents_count, deployments_count = fbb_importer.sync_data()
        return jsonify({
            "message": "Sincronización exitosa desde Google Sheets",
            "zones_imported": zones_count,
            "boxes_imported": boxes_count,
            "staff_imported": staff_count,
            "incidents_imported": incidents_count,
            "deployments_imported": deployments_count
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff/export", methods=["GET"])
def fbb_export_staff():
    try:
        csv_data = FBBManager.export_staff_csv()
        csv_bytes = csv_data.encode("utf-8-sig")
        from flask import Response
        return Response(
            csv_bytes,
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=plantilla_staff.csv"}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/staff/import", methods=["POST"])
def fbb_import_staff():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No se subió ningún archivo."}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Nombre de archivo vacío."}), 400
        if not file.filename.endswith('.csv'):
            return jsonify({"error": "El archivo debe ser un formato CSV."}), 400
        csv_content = file.read()
        created, updated, errors = FBBManager.import_staff_csv(csv_content)
        return jsonify({
            "message": "Importación procesada.",
            "created": created,
            "updated": updated,
            "errors": errors
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/partners/capacity", methods=["GET"])
def fbb_get_partners_capacity_report():
    try:
        report = FBBManager.get_partner_capacity_report()
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/zones/capacity-detail", methods=["GET"])
def fbb_get_zones_capacity_detail_report():
    try:
        detail = FBBManager.get_zone_capacity_detail()
        return jsonify(detail)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/charts/branch-capacity-stacked", methods=["GET"])
def fbb_get_branch_capacity_stacked():
    try:
        branch = request.args.get("branch", "").strip()
        partner = request.args.get("partner", "").strip()
        zone = request.args.get("zone", "").strip()
        report = FBBManager.get_branch_capacity_stacked_report(branch=branch, partner=partner, zone=zone)
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/stats", methods=["GET"])
def fbb_get_incidents_stats():
    try:
        site = request.args.get("site", "").strip()
        branch = request.args.get("branch", "").strip()
        month = request.args.get("month", "").strip()
        week = request.args.get("week", "").strip()

        evolution = FBBManager.get_incidents_by_month(branch=branch, month=month, site=site, week=week)
        status = FBBManager.get_incidents_by_status(branch=branch, month=month, site=site, week=week)
        ranking = FBBManager.get_incidents_sites_ranking(branch=branch, month=month, site=site, week=week)
        monthly_breakdown = FBBManager.get_incidents_monthly_breakdown(branch=branch, month=month, site=site, week=week)
        site_breakdown = FBBManager.get_incidents_site_breakdown(branch=branch, month=month, site=site, week=week)

        conn = FBBManager.get_connection()
        cursor = conn.cursor()
        query_parts = ["WHERE 1=1"]
        params = []
        if branch:
            query_parts.append("branch = ?")
            params.append(branch)
        if week:
            query_parts.append("week_number = ?")
            params.append(week)
        elif month:
            query_parts.append("month_year = ?")
            params.append(month)
        if site:
            query_parts.append("station_code = ?")
            params.append(site)
        where_clause = " AND ".join(query_parts)
        cursor.execute(f"SELECT COUNT(*), COUNT(DISTINCT subscriber) FROM incidents {where_clause}", params)
        counts = cursor.fetchone()
        tot_inc = counts[0] or 0
        uniq_cli = counts[1] or 0
        conn.close()

        most_freq_status = status[0]["status_desc"] if status else "N/A"
        most_freq_count = status[0]["count"] if status else 0

        return jsonify({
            "evolution": evolution,
            "status": status,
            "ranking": ranking,
            "monthly_breakdown": monthly_breakdown,
            "site_breakdown": site_breakdown,
            "kpis": {
                "total_incidents": tot_inc,
                "unique_clients": uniq_cli,
                "most_frequent_status": most_freq_status,
                "most_frequent_count": most_freq_count
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/sites", methods=["GET"])
def fbb_get_incidents_sites():
    try:
        branch = request.args.get("branch", "").strip()
        sites = FBBManager.get_incidents_sites(branch=branch)
        return jsonify(sites)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/months", methods=["GET"])
def fbb_get_incidents_months():
    try:
        months = FBBManager.get_incidents_months()
        return jsonify(months)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/deployments/months", methods=["GET"])
def fbb_get_deployments_months():
    try:
        months = FBBManager.get_deployments_months()
        return jsonify(months)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/outages", methods=["GET"])
def fbb_get_site_outages():
    try:
        branch = request.args.get("branch", "").strip()
        month = request.args.get("month", "").strip()
        site = request.args.get("site", "").strip()
        week = request.args.get("week", "").strip()

        outages_ranking = FBBManager.get_site_outages_report(branch=branch, month=month, site=site, week=week)
        causes = FBBManager.get_site_outage_causes(branch=branch, month=month, site=site, week=week)

        total_energy_cuts = sum(r.get("energy_cuts", 0) for r in outages_ranking)
        total_energy_affected = sum(r.get("energy_affected", 0) for r in outages_ranking)
        total_odn_cuts = sum(r.get("odn_cuts", 0) for r in outages_ranking)
        total_odn_affected = sum(r.get("odn_affected", 0) for r in outages_ranking)
        total_wos = sum(r.get("total_wos", 0) for r in outages_ranking)

        return jsonify({
            "outages_ranking": outages_ranking,
            "causes": causes,
            "kpis": {
                "total_energy_cuts": total_energy_cuts,
                "total_energy_affected": total_energy_affected,
                "total_odn_cuts": total_odn_cuts,
                "total_odn_affected": total_odn_affected,
                "total_wos": total_wos
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/outages/details", methods=["GET"])
def fbb_get_site_outage_details():
    try:
        site = request.args.get("site", "").strip()
        branch = request.args.get("branch", "").strip()
        month = request.args.get("month", "").strip()
        week = request.args.get("week", "").strip()
        if not site:
            return jsonify({"error": "site parameter is required"}), 400
        details = FBBManager.get_site_outage_details(site=site, branch=branch, month=month, week=week)
        return jsonify(details)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/wos/details", methods=["GET"])
def fbb_get_site_wo_details():
    try:
        site = request.args.get("site", "").strip()
        branch = request.args.get("branch", "").strip()
        month = request.args.get("month", "").strip()
        week = request.args.get("week", "").strip()
        if not site:
            return jsonify({"error": "site parameter is required"}), 400
        details = FBBManager.get_site_wo_details(site=site, branch=branch, month=month, week=week)
        return jsonify(details)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/incidents/weeks", methods=["GET"])
def fbb_get_incidents_weeks():
    try:
        weeks = FBBManager.get_incidents_weeks()
        return jsonify(weeks)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/fbb/deployments/stats", methods=["GET"])
def fbb_get_deployments_stats():
    try:
        branch = request.args.get("branch", "").strip()
        month = request.args.get("month", "").strip()
        stats = FBBManager.get_deployments_report(branch=branch, month=month)
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/fbb-static/<path:path>")
def fbb_static_proxy(path):
    return send_from_directory(os.path.join(app.static_folder, "..", "DataBaseFBB", "static"), path)


# Servir archivos estáticos secundarios (JS, CSS, imágenes)
# ESTE DEBE IR AL FINAL PARA EVITAR QUE SE COMA OTRAS RUTAS COMO /api
@app.route("/<path:path>")
def static_proxy(path):
    return send_from_directory(app.static_folder, path)

AUTO_SYNC_LOOP_INTERVAL_SECONDS = 15 * 60  # pausa entre el fin de un ciclo y el inicio del siguiente
AUTO_DEPLOY_LOOP_INTERVAL_SECONDS = 15 * 60

def _auto_excel_sync_loop():
    """Reemplaza a la Tarea Programada de Windows (FBB_GNOC_FullSync, ahora desactivada):
    mientras este servidor esté corriendo, sincroniza Excel y sube a Google Sheets en bucle
    continuo, sin depender de que el Programador de Tareas esté bien configurado. Reusa
    exactamente la misma run_background_sync() que dispara el botón "Sincronizar Excel" -si
    el usuario ya está sincronizando a mano cuando le toca a este ciclo, se salta esa vuelta
    en vez de pelear por el mismo sync_lock."""
    while True:
        try:
            with sync_lock:
                ya_corriendo = sync_status["state"] in ("downloading", "processing")
            if ya_corriendo:
                print("[AutoSync] Ya hay una sincronización en curso (probablemente manual) -- se omite este ciclo.", flush=True)
            else:
                gnoc_env_overrides = None
                from_month, to_month = load_sync_month_range()
                if from_month and to_month:
                    gnoc_env_overrides = {"FILTER_CREATE_TIME": compute_filter_create_time_for_months(from_month, to_month)}
                    print(f"[AutoSync] Usando el último rango de meses elegido en el dashboard: {from_month} a {to_month}", flush=True)
                print("[AutoSync] Iniciando sincronización automática de Excel...", flush=True)
                run_background_sync(gnoc_env_overrides)
                with sync_lock:
                    estado_final = dict(sync_status)
                if estado_final["state"] == "success":
                    print(f"[AutoSync] Sync OK: {estado_final['message']} -- subiendo a Google Sheets...", flush=True)
                    try:
                        filas = sheets_push.push_pending_valid_to_sheet()
                        print(f"[AutoSync] Push a Google Sheets OK ({filas} filas).", flush=True)
                    except Exception as e:
                        print(f"[AutoSync] Push a Google Sheets falló: {e}", flush=True)
                    try:
                        errores = sheets_push.push_marlo_errors_to_sheet()
                        print(f"[AutoSync] Push de errores de Marlo a ERRORES LVL3 OK ({errores} nuevos).", flush=True)
                    except Exception as e:
                        print(f"[AutoSync] Push de errores de Marlo a ERRORES LVL3 falló: {e}", flush=True)
                    try:
                        nuevas_gnocall = sheets_push.push_gnocall_to_sheet()
                        print(f"[AutoSync] Push a GNOCALL OK ({nuevas_gnocall} nuevas).", flush=True)
                    except Exception as e:
                        print(f"[AutoSync] Push a GNOCALL falló: {e}", flush=True)
                    if cloud_sync.is_configured():
                        try:
                            payload = daily_report.build_cloud_payload()
                            result = cloud_sync.push_daily_report(payload)
                            print(f"[AutoSync] Push Reporte Diario a la nube: {result}", flush=True)
                        except Exception as e:
                            print(f"[AutoSync] Push de Reporte Diario a la nube falló: {e}", flush=True)
                else:
                    print(f"[AutoSync] Sync de Excel terminó en estado '{estado_final['state']}' -- se omite el push a Sheets.", flush=True)
        except Exception as e:
            print(f"[AutoSync] Error inesperado en el ciclo automático: {e}", flush=True)
        # Siempre espera el intervalo COMPLETO tras terminar (sin restar cuánto duró el
        # ciclo) -- si se restaba, un ciclo que tardaba más de 15 min (pasa seguido con
        # GNOC/Tableau) dejaba la pausa en 0 y arrancaba el siguiente de inmediato, lo que
        # causó procesos zombis compitiendo por los mismos recursos (ver incidente de hoy).
        print(f"[AutoSync] Esperando {AUTO_SYNC_LOOP_INTERVAL_SECONDS // 60} min antes del siguiente ciclo...", flush=True)
        time.sleep(AUTO_SYNC_LOOP_INTERVAL_SECONDS)

def _auto_deploy_pending_loop():
    """Igual que _auto_excel_sync_loop pero para "Actualizar Despliegues" (deploy ant):
    trigger_run() ya trae su propio guard contra corridas simultáneas (devuelve False si
    ya hay una activa, manual o automática), así que alcanza con revisar ese resultado."""
    while True:
        try:
            iniciado = deploy_pending.trigger_run()
            if not iniciado:
                print("[AutoDeploy] Ya hay una actualización de despliegues en curso -- se omite este ciclo.", flush=True)
            else:
                print("[AutoDeploy] Iniciando actualización automática de despliegues...", flush=True)
                while deploy_pending.get_run_status()["state"] == "running":
                    time.sleep(5)
                estado_final = deploy_pending.get_run_status()
                print(f"[AutoDeploy] Terminó en estado '{estado_final['state']}': {estado_final['message']}", flush=True)
        except Exception as e:
            print(f"[AutoDeploy] Error inesperado en el ciclo automático: {e}", flush=True)
        print(f"[AutoDeploy] Esperando {AUTO_DEPLOY_LOOP_INTERVAL_SECONDS // 60} min antes del siguiente ciclo...", flush=True)
        time.sleep(AUTO_DEPLOY_LOOP_INTERVAL_SECONDS)

if __name__ == "__main__":
    print("Iniciando servidor de la Plataforma de Avance GNOC en http://localhost:5001 ...")
    threading.Thread(target=_alarm_scan_loop, daemon=True).start()
    threading.Thread(target=_cloud_comments_pull_loop, daemon=True).start()
    threading.Thread(target=_auto_excel_sync_loop, daemon=True).start()
    threading.Thread(target=_auto_deploy_pending_loop, daemon=True).start()
    # Loop automático de Auditoría OLT — arranca siempre; si falta olts_input.xlsx
    # el loop espera 60 s y reintenta hasta que el archivo aparezca.
    threading.Thread(target=_olt_scan_loop, daemon=True).start()
    # threaded=True: las consultas EN VIVO de puerto OLT (get_olt_port_status_live) hacen
    # una llamada SOAP bloqueante de hasta unos segundos (más si hay reintentos); sin esto
    # el servidor de desarrollo de Flask atiende un solo request a la vez y esa consulta
    # dejaría "colgado" el resto del dashboard mientras espera al EMS.
    app.run(host="0.0.0.0", port=5001, debug=False, threaded=True)
