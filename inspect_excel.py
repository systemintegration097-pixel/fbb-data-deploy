import openpyxl

wb = openpyxl.load_workbook("reporte_gnoc.xlsx", data_only=True)
print("Hojas del libro:", wb.sheetnames)
sheet = wb.active
print(f"Dimensiones de la hoja activa: {sheet.dimensions}")
print(f"max_row: {sheet.max_row}, max_column: {sheet.max_column}")

print("\nPrimeras 15 filas:")
for r in range(1, min(16, sheet.max_row + 1)):
    row_vals = [cell.value for cell in sheet[r]]
    print(f"Fila {r}: {row_vals}")

wb.close()
